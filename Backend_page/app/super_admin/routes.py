from flask import Blueprint, jsonify,request
#from app.middleware.role_required import role_required
import json
import os
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
from app.init_db import get_db_connection
from mysql.connector import Error

from flask import Blueprint, jsonify, request, send_file, make_response, send_from_directory, after_this_request
import json
import os
import pandas as pd
from io import BytesIO
from datetime import datetime, date, timedelta
from app.middleware.role_required import role_required
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
import openpyxl.styles
import openpyxl.utils
super_admin_bp = Blueprint("super_admin", __name__)

#@role_required("super_admin")

@super_admin_bp.route("/api/events_detail", methods=["GET"])
def events():

    db = get_db_connection()
    cursor = db.cursor()

    try:

        cursor.execute("""
            SELECT id, event_code, event_name 
            FROM event_details_table
        """)

        rows = cursor.fetchall()
        print("Event",rows)

        events = []

        for row in rows:
            events.append({
                "id": row[0],
                "event_code": row[1],
                "event_name": row[2]
            })

        return jsonify({
            "status": "success",
            "data": events
        })

    except Exception as e:

        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    finally:
        cursor.close()
        db.close()

@super_admin_bp.route("/api/live_food_count", methods=["POST"])
def get_live_food_count():

    db = get_db_connection()
    cursor = db.cursor()

    try:

        data = request.json

        event_id = data.get("event_id")
        meal_time = data.get("meal_time")
        meal_type = data.get("meal_type")

        cursor.execute("""
            SELECT guests_inside, total_capacity, waiting_outside
            FROM food_live_count
            WHERE event_id=%s AND meal_time=%s AND meal_type=%s
        """,(event_id, meal_time, meal_type))

        row = cursor.fetchone()
        print("Food",row)

        if row:

            result = {
                "guests_inside": row[0],
                "total_capacity": row[1],
                "waiting_outside": row[2]
            }

        else:

            result = {
                "guests_inside": 0,
                "total_capacity": 0,
                "waiting_outside": 0
            }

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:

        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    finally:
        cursor.close()
        db.close()

# -----------------------------------
# GET COUNTRIES
# -----------------------------------

@super_admin_bp.route("/api/countries", methods=["GET"])
def get_countries():

    try:

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM countries")

        data = cursor.fetchall()

        return jsonify(data)

    except Exception as e:

        return jsonify({"error": str(e)})

    finally:

        cursor.close()
        conn.close()


# -----------------------------------
# GET STATES
# -----------------------------------

@super_admin_bp.route("/api/states/<country_id>", methods=["GET"])
def get_states(country_id):

    try:

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            "SELECT * FROM states WHERE country_id=%s",
            (country_id,)
        )

        data = cursor.fetchall()

        return jsonify(data)

    except Exception as e:

        return jsonify({"error": str(e)})

    finally:

        cursor.close()
        conn.close()


# -----------------------------------
# GET CITIES
# -----------------------------------

@super_admin_bp.route("/api/cities/<state_id>", methods=["GET"])
def get_cities(state_id):

    try:

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            "SELECT * FROM cities WHERE state_id=%s",
            (state_id,)
        )

        data = cursor.fetchall()

        return jsonify(data)

    except Exception as e:

        return jsonify({"error": str(e)})

    finally:

        cursor.close()
        conn.close()


# -----------------------------------
# CREATE VENUE
# -----------------------------------
@super_admin_bp.route("/api/create_venue", methods=["POST"])
def create_venue():

    try:

        data = request.json
        print("Create", data)

        venue_name = data.get("venue_name")
        address = data.get("address")
        country_id = data.get("country")
        state_id = data.get("state")
        city_id = data.get("city")
        pin_code = data.get("pin_code")
        status = data.get("status")
        venue_image = data.get("venue_image")

        documents = data.get("documents")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # get country name
        if str(country_id).isdigit():
            cursor.execute("SELECT country_name FROM countries WHERE id=%s", (country_id,))
            country_row = cursor.fetchone()
            country = country_row["country_name"] if country_row else country_id
        else:
            country = country_id

        # get state name
        if str(state_id).isdigit():
            cursor.execute("SELECT state_name FROM states WHERE id=%s", (state_id,))
            state_row = cursor.fetchone()
            state = state_row["state_name"] if state_row else state_id
        else:
            state = state_id

        # get city name
        if str(city_id).isdigit():
            cursor.execute("SELECT city_name FROM cities WHERE id=%s", (city_id,))
            city_row = cursor.fetchone()
            city = city_row["city_name"] if city_row else city_id
        else:
            city = city_id

        # generate venue code
        cursor.execute("SELECT MAX(id) as max_id FROM venues")
        result = cursor.fetchone()

        new_number = (result["max_id"] or 0) + 1
        venue_code = f"VEN-{new_number}"

        organizer_id = data.get("organizer_id")
        created_by = data.get("created_by", "System")
        
        if organizer_id and created_by == "System":
            cursor.execute("SELECT name FROM users WHERE id = %s", (organizer_id,))
            user_row = cursor.fetchone()
            if user_row:
                created_by = user_row["name"]

        # insert venue
        cursor.execute(
            """
            INSERT INTO venues
            (venue_code,venue_name,address,country_name,state_name,city_name,pin_code,venue_image,status,organizer_id,created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                venue_code,
                venue_name,
                address,
                country,
                state,
                city,
                pin_code,
                venue_image,
                status,
                organizer_id,
                created_by
            )
        )

        venue_id = cursor.lastrowid

        # insert documents
        if documents:

            for doc in documents:

                cursor.execute(
                    """
                    INSERT INTO venue_documents
                    (venue_id,document_type,document_number,document_file)
                    VALUES (%s,%s,%s,%s)
                    """,
                    (
                        venue_id,
                        doc["document_type"],
                        doc["document_number"],
                        doc["document_file"]
                    )
                )

        conn.commit()

        return jsonify({
            "message": "Venue Created",
            "venue_code": venue_code
        })

    except Exception as e:
        print("Create Venue Error:", str(e))
        return jsonify({
            "error": str(e)
        }), 500

    finally:

        cursor.close()
        conn.close()

# -----------------------------------
# UPDATE VENUE
# -----------------------------------
@super_admin_bp.route("/api/update_venue/<int:id>", methods=["PUT"])
def update_venue(id):
    try:
        data = request.json
        venue_name = data.get("venue_name")
        address = data.get("address")
        country_val = data.get("country")
        state_val = data.get("state")
        city_val = data.get("city")
        pin_code = data.get("pin_code")
        status = data.get("status")
        venue_image = data.get("venue_image")
        documents = data.get("documents")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        if str(country_val).isdigit():
            cursor.execute("SELECT country_name FROM countries WHERE id=%s", (country_val,))
            country_row = cursor.fetchone()
            country = country_row["country_name"] if country_row else ""
        else:
            country = country_val

        if str(state_val).isdigit():
            cursor.execute("SELECT state_name FROM states WHERE id=%s", (state_val,))
            state_row = cursor.fetchone()
            state = state_row["state_name"] if state_row else ""
        else:
            state = state_val

        if str(city_val).isdigit():
            cursor.execute("SELECT city_name FROM cities WHERE id=%s", (city_val,))
            city_row = cursor.fetchone()
            city = city_row["city_name"] if city_row else ""
        else:
            city = city_val

        organizer_id = data.get("organizer_id")
        modified_by = data.get("modified_by", "System")
        
        if organizer_id and modified_by == "System":
            cursor.execute("SELECT name FROM users WHERE id = %s", (organizer_id,))
            user_row = cursor.fetchone()
            if user_row:
                modified_by = user_row["name"]

        cursor.execute(
            """
            UPDATE venues
            SET venue_name=%s, address=%s, country_name=%s, state_name=%s, city_name=%s, 
                pin_code=%s, venue_image=%s, status=%s, modified_by=%s
            WHERE id=%s
            """,
            (venue_name, address, country, state, city, pin_code, venue_image, status, modified_by, id)
        )

        cursor.execute("DELETE FROM venue_documents WHERE venue_id=%s", (id,))
        if documents:
            for doc in documents:
                cursor.execute(
                    """
                    INSERT INTO venue_documents (venue_id, document_type, document_number, document_file)
                    VALUES (%s, %s, %s, %s)
                    """,
                    (id, doc.get("document_type"), doc.get("document_number"), doc.get("document_file"))
                )

        conn.commit()

        return jsonify({"message": "Venue Updated"})

    except Exception as e:
        print("Update Venue Error:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()

# -----------------------------------
# GET VENUE LIST
# -----------------------------------

@super_admin_bp.route("/api/venues", methods=["GET"])
def get_venues():

    try:
        organizer_id = request.args.get("organizer_id")
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        if organizer_id:
            query = "SELECT id,venue_code,venue_name,address,status,created_by,created_at as created_on,modified_by,modified_on FROM venues WHERE organizer_id = %s"
            cursor.execute(query, (organizer_id,))
        else:
            query = "SELECT id,venue_code,venue_name,address,status,created_by,created_at as created_on,modified_by,modified_on FROM venues"
            cursor.execute(query)

        data = cursor.fetchall()
        print(data)

        return jsonify(data)

    except Exception as e:

        return jsonify({"error": str(e)})

    finally:

        cursor.close()
        conn.close()

@super_admin_bp.route("/api/delete_venue/<int:id>", methods=["DELETE"])
def delete_venue(id):
    try:
        print("Venue",id)
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # ✅ Check if venue exists
        cursor.execute("SELECT * FROM venues WHERE id = %s", (id,))
        venue = cursor.fetchone()

        if not venue:
            return jsonify({
                "status": False,
                "message": "Venue not found"
            }), 404

        # ✅ Delete associated documents first
        cursor.execute("DELETE FROM venue_documents WHERE venue_id = %s", (id,))

        # ✅ Delete venue
        cursor.execute("DELETE FROM venues WHERE id = %s", (id,))
        conn.commit()

        return jsonify({
            "status": True,
            "message": "Venue deleted successfully"
        }), 200

    except Exception as e:
        print("ERROR:", str(e))
        return jsonify({
            "status": False,
            "message": "Failed to delete venue",
            "error": str(e)
        }), 500

    finally:
        cursor.close()
        conn.close()
# -----------------------------------
# VIEW VENUE DETAILS
# -----------------------------------

@super_admin_bp.route("/api/venuedetail/<id>", methods=["GET"])
def venue_details(id):

    try:

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            "SELECT * FROM venues WHERE id=%s",
            (id,)
        )

        venue = cursor.fetchone()
        if venue and venue.get("venue_image"):
            venue["venue_image"] = get_image_url(venue["venue_image"])

        cursor.execute(
            "SELECT * FROM venue_documents WHERE venue_id=%s",
            (id,)
        )

        docs = cursor.fetchall()
        for d in docs:
            if d.get("document_file"):
                d["document_file"] = get_image_url(d["document_file"])

        return jsonify({
            "venue": venue,
            "documents": docs
        })

    except Exception as e:

        return jsonify({"error": str(e)})

    finally:

        cursor.close()
        conn.close()


@super_admin_bp.route("/api/create_sponsor", methods=["POST"])
def create_sponsor_detail():

    conn = None
    cursor = None

    try:

        data = request.json
        print("Data:", data)

        sponsor_name = data["sponsor_name"]
        primary_contact = data["primary_contact"]
        secondary_contact = data.get("secondary_contact")
        mail_id = data["mail_id"]
        address = data["address"]
        status = data.get("status", "Active")
        organizer_id = data.get("organizer_id")

        conn = get_db_connection()
        cursor = conn.cursor()

        # Generate sponsor code
        cursor.execute("SELECT COUNT(*) FROM sponsors_details")
        count = cursor.fetchone()[0] + 1
        sponsor_code = f"SP{str(count).zfill(4)}"

        created_by = data.get("created_by", "System")
        if organizer_id and created_by == "System":
            cursor.execute("SELECT name FROM users WHERE id = %s", (organizer_id,))
            user_row = cursor.fetchone()
            if user_row:
                created_by = user_row[0]

        cursor.execute(
            """
            INSERT INTO sponsors_details 
            (sponsor_code, sponsor_name, primary_contact, secondary_contact, mail_id, address, status, organizer_id, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                sponsor_code,
                sponsor_name,
                primary_contact,
                secondary_contact,
                mail_id,
                address,
                status,
                organizer_id,
                created_by
            )
        )

        sponsor_id = cursor.lastrowid

        # Insert Documents
        documents = data.get("documents", [])
        for doc in documents:
            cursor.execute("""
                INSERT INTO sponsor_documents
                (sponsor_id, document_type, document_number, document_file)
                VALUES (%s, %s, %s, %s)
            """, (
                sponsor_id,
                doc.get("document_type"),
                doc.get("document_number"),
                doc.get("document_file")
            ))

        conn.commit()

        return jsonify({
            "status": "success",
            "message": "Sponsor created successfully",
            "sponsor_id": sponsor_id
        }), 201

    except Exception as e:

        print("ERROR:", str(e))

        if conn:
            conn.rollback()

        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

@super_admin_bp.route("/api/sponsors", methods=["GET"])
def get_sponsors():
    organizer_id = request.args.get("organizer_id")
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if organizer_id:
        cursor.execute("""
            SELECT id, sponsor_code, sponsor_name, primary_contact, secondary_contact, mail_id, address, status, 
                   created_by, created_at as created_on, modified_by, modified_on 
            FROM sponsors_details 
            WHERE organizer_id = %s 
            ORDER BY id DESC
        """, (organizer_id,))
    else:
        cursor.execute("""
            SELECT id, sponsor_code, sponsor_name, primary_contact, secondary_contact, mail_id, address, status, 
                   created_by, created_at as created_on, modified_by, modified_on 
            FROM sponsors_details 
            ORDER BY id DESC
        """)

    sponsors = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(sponsors)

@super_admin_bp.route("/api/sponsor/<int:id>", methods=["GET"])
def view_sponsor(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM sponsors_details WHERE id=%s", (id,))
    sponsor = cursor.fetchone()
    
    if sponsor:
        cursor.execute("SELECT * FROM sponsor_documents WHERE sponsor_id=%s", (id,))
        docs = cursor.fetchall()
        for d in docs:
            if d.get("document_file"):
                d["document_file"] = get_image_url(d["document_file"])
        sponsor["documents"] = docs
        
    cursor.close()
    conn.close()
    return jsonify(sponsor)

@super_admin_bp.route("/api/update_sponsor/<int:id>", methods=["PUT"])
def update_sponsor(id):
    try:
        data = request.json
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        organizer_id = data.get("organizer_id")
        modified_by = data.get("modified_by", "System")
        
        if organizer_id and modified_by == "System":
            cursor.execute("SELECT name FROM users WHERE id = %s", (organizer_id,))
            user_row = cursor.fetchone()
            if user_row:
                modified_by = user_row["name"]

        cursor.execute("""
            UPDATE sponsors_details 
            SET sponsor_name=%s, primary_contact=%s, secondary_contact=%s, mail_id=%s, address=%s, status=%s, modified_by=%s
            WHERE id=%s
        """, (
            data.get("sponsor_name"),
            data.get("primary_contact"),
            data.get("secondary_contact"),
            data.get("mail_id"),
            data.get("address"),
            data.get("status", "Active"),
            modified_by,
            id
        ))

        # Update documents
        cursor.execute("DELETE FROM sponsor_documents WHERE sponsor_id = %s", (id,))
        documents = data.get("documents", [])
        for doc in documents:
            cursor.execute("""
                INSERT INTO sponsor_documents
                (sponsor_id, document_type, document_number, document_file)
                VALUES (%s, %s, %s, %s)
            """, (
                id,
                doc.get("document_type"),
                doc.get("document_number"),
                doc.get("document_file")
            ))

        conn.commit()
        return jsonify({"status": "success", "message": "Sponsor updated successfully"}), 200

    except Exception as e:
        print("Update Sponsor Error:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()
@super_admin_bp.route("/api/delete_sponsor/<int:id>", methods=["DELETE"])
def delete_sponsor(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Check if sponsor exists
        cursor.execute("SELECT * FROM sponsors_details WHERE id = %s", (id,))
        sponsor = cursor.fetchone()

        if not sponsor:
            return jsonify({
                "status": False,
                "message": "Sponsor not found"
            }), 404

        # Delete documents first
        cursor.execute("DELETE FROM sponsor_documents WHERE sponsor_id = %s", (id,))
        
        # Delete sponsor
        cursor.execute("DELETE FROM sponsors_details WHERE id = %s", (id,))
        conn.commit()

        return jsonify({
            "status": True,
            "message": "Sponsor deleted successfully"
        }), 200

    except Exception as e:
        print("ERROR:", str(e))
        return jsonify({
            "status": False,
            "message": "Failed to delete sponsor",
            "error": str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# ================= SPONSOR EXPORT =================

@super_admin_bp.route("/api/sponsors/export/excel", methods=["GET"])
def export_sponsors_excel():
    try:
        print(f"INFO: [{datetime.now()}] Initiating Excel export for Sponsor Management.")
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
        SELECT 
            sponsor_code,
            sponsor_name,
            primary_contact,
            mail_id,
            address,
            status,
            created_by,
            created_at as created_on,
            modified_by,
            modified_on
        FROM sponsors_details ORDER BY id DESC
        """)
        data = cursor.fetchall()
        
        if not data:
            return jsonify({"error": "No data available for export"}), 404

        df = pd.DataFrame(data)
        df["created_on"] = pd.to_datetime(df["created_on"]).dt.tz_localize(None)
        df["modified_on"] = pd.to_datetime(df["modified_on"]).dt.tz_localize(None)
        df.columns = [
            "Sponsor Code",
            "Sponsor Name",
            "Contact",
            "Email",
            "Address",
            "Status",
            "Created By",
            "Created On",
            "Modified By",
            "Modified On"
        ]
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sponsor_Report')
            workbook = writer.book
            worksheet = writer.sheets['Sponsor_Report']

            # Apply date format
            for cell in worksheet["H"][1:]:
                cell.number_format = 'DD/MM/YYYY'
            for cell in worksheet["J"][1:]:
                cell.number_format = 'DD/MM/YYYY'

            # Header Styling
            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")

            output.seek(0)
            filename = f"sponsor_list_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
        
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response

    except Exception as e:
        print(f"ERROR: [{datetime.now()}] Excel export failed: {str(e)}")
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@super_admin_bp.route("/api/sponsors/export/pdf", methods=["GET"])
def export_sponsors_pdf():
    try:
        print(f"INFO: [{datetime.now()}] Initiating PDF export for Sponsor Management.")
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                sponsor_name,
                primary_contact,
                status,
                created_by,
                DATE_FORMAT(created_at, '%d-%m-%Y') as created_on,
                modified_by,
                DATE_FORMAT(modified_on, '%d-%m-%Y') as modified_on
            FROM sponsors_details 
            ORDER BY id DESC
        """)
        data = cursor.fetchall()

        if not data:
            return jsonify({"error": "No data available for export"}), 404

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = styles['Title']
        title_style.fontSize = 20
        title_style.textColor = colors.HexColor("#1e40af")
        elements.append(Paragraph("Sponsor Management Report", title_style))
        elements.append(Spacer(1, 24))

        table_data = [["Sponsor Name", "Contact", "Status", "Created By", "Created On", "Modified By", "Modified On"]]
        for row in data:
            table_data.append([
                row['sponsor_name'],
                row['primary_contact'],
                row['status'],
                row['created_by'] or 'System',
                row['created_on'] or '',
                row['modified_by'] or 'N/A',
                row['modified_on'] or 'N/A'
            ])

        column_widths = [140, 90, 70, 100, 90, 100, 90]
        t = Table(table_data, colWidths=column_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        
        elements.append(t)
        doc.build(elements)
        output.seek(0)
        filename = f"sponsor_list_{datetime.now().strftime('%Y_%m_%d')}.pdf"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')

    except Exception as e:
        print(f"ERROR: [{datetime.now()}] PDF export failed: {str(e)}")
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@super_admin_bp.route("/api/create_vendor", methods=["POST"])
def create_vendor():

    conn = None
    cursor = None

    try:

        data = request.json
        print("Data",data)

        vendor_type = data["vendor_type"]
        vendor_name = data["vendor_name"]
        company_name = data["company_name"]
        primary_contact = data["primary_contact"]
        secondary_contact = data.get("secondary_contact")
        mail_id = data["mail_id"]
        address = data["address"]
        country = data.get("country")
        state = data.get("state")
        city = data.get("city")

        bank_name = data.get("bank_name")
        account_holder = data.get("account_holder")
        ifsc_code = data.get("ifsc_code")
        account_number = data.get("account_number")

        status = data.get("status", "Active")
        organizer_id = data.get("organizer_id")
        bank_passbook = data.get("bank_passbook")
        created_by = data.get("created_by", "System")

        documents = data.get("documents", [])

        conn = get_db_connection()
        cursor = conn.cursor()

        if organizer_id and created_by == "System":
            cursor.execute("SELECT name FROM users WHERE id = %s", (organizer_id,))
            user_row = cursor.fetchone()
            if user_row:
                created_by = user_row[0]

        cursor.execute("""
        INSERT INTO vendor_details
        (vendor_type,vendor_name,company_name,primary_contact,secondary_contact,mail_id,country,state,city,address,
        bank_name,account_holder,ifsc_code,account_number,bank_passbook,status,organizer_id,created_by)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,(
            vendor_type,
            vendor_name,
            company_name,
            primary_contact,
            secondary_contact,
            mail_id,
            country,
            state,
            city,
            address,
            bank_name,
            account_holder,
            ifsc_code,
            account_number,
            bank_passbook,
            status,
            organizer_id,
            created_by
        ))

        vendor_id = cursor.lastrowid

        # Insert Documents

        for doc in documents:

            cursor.execute("""
            INSERT INTO vendor_documents
            (vendor_id,document_type,document_number,document_file)
            VALUES (%s,%s,%s,%s)
            """,(
                vendor_id,
                doc["document_type"],
                doc["document_number"],
                doc["document_file"]
            ))

        conn.commit()

        return jsonify({
            "status":"success",
            "message":"Vendor created successfully"
        })

    except Exception as e:

        if conn:
            conn.rollback()

        return jsonify({
            "status":"error",
            "message":str(e)
        }),500

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


#new Upaded Code on 15-05-2026
@super_admin_bp.route("/api/update_vendor/<int:id>", methods=["PUT"])
def update_vendor(id):

    conn = None
    cursor = None

    try:

        data = request.json
        print("Update Data", data)

        vendor_type = data.get("vendor_type")
        vendor_name = data.get("vendor_name")
        company_name = data.get("company_name")
        primary_contact = data.get("primary_contact")
        secondary_contact = data.get("secondary_contact")
        mail_id = data.get("mail_id")
        country = data.get("country")
        state = data.get("state")
        city = data.get("city")
        address = data.get("address")

        bank_name = data.get("bank_name")
        account_holder = data.get("account_holder")
        ifsc_code = data.get("ifsc_code")
        account_number = data.get("account_number")
        bank_passbook = data.get("bank_passbook")

        status = data.get("status", "Active")
        organizer_id = data.get("organizer_id")
        modified_by = data.get("modified_by", "System")

        documents = data.get("documents", [])

        conn = get_db_connection()
        cursor = conn.cursor()

        if organizer_id and modified_by == "System":
            cursor.execute("SELECT name FROM users WHERE id = %s", (organizer_id,))
            user_row = cursor.fetchone()
            if user_row:
                modified_by = user_row[0]

        cursor.execute("""
        UPDATE vendor_details SET 
        vendor_type=%s, vendor_name=%s, company_name=%s, primary_contact=%s, 
        secondary_contact=%s, mail_id=%s, country=%s, state=%s, city=%s, address=%s, bank_name=%s, 
        account_holder=%s, ifsc_code=%s, account_number=%s, bank_passbook=%s, status=%s, organizer_id=COALESCE(%s, organizer_id), modified_by=%s
        WHERE id=%s
        """, (
            vendor_type, vendor_name, company_name, primary_contact,
            secondary_contact, mail_id, country, state, city, address, bank_name,
            account_holder, ifsc_code, account_number, bank_passbook, status, organizer_id, modified_by, id
        ))

        # Delete existing documents
        cursor.execute("DELETE FROM vendor_documents WHERE vendor_id=%s", (id,))

        # Insert Documents
        for doc in documents:
            cursor.execute("""
            INSERT INTO vendor_documents
            (vendor_id,document_type,document_number,document_file)
            VALUES (%s,%s,%s,%s)
            """, (
                id,
                doc["document_type"],
                doc["document_number"],
                doc["document_file"]
            ))

        conn.commit()

        return jsonify({
            "status": "success",
            "message": "Vendor updated successfully"
        })

    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# ================= VENDOR EXPORT =================
#new Upaded Code on 15-05-2026
@super_admin_bp.route("/api/vendors/export/excel", methods=["GET"])
def export_vendors_excel():
    try:
        print(f"INFO: [{datetime.now()}] Initiating Excel export for Vendor Management.")
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
        SELECT 
            vendor_name,
            primary_contact,
            mail_id,
            country,
            state,
            city,
            address,
            status,
            created_by,
            created_at as created_on,
            modified_by,
            modified_on
        FROM vendor_details ORDER BY id DESC
        """)
        data = cursor.fetchall()
        
        if not data:
            return jsonify({"error": "No data available for export"}), 404

        df = pd.DataFrame(data)
        df["created_on"] = pd.to_datetime(df["created_on"]).dt.tz_localize(None)
        df["modified_on"] = pd.to_datetime(df["modified_on"]).dt.tz_localize(None)
        df.columns = [
            "Vendor Name",
            "Contact",
            "Email",
            "Country",
            "State",
            "City",
            "Address",
            "Status",
            "Created By",
            "Created On",
            "Modified By",
            "Modified On"
        ]
        
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            df.to_excel(writer, index=False, sheet_name='Vendor_Report')

            workbook = writer.book
            worksheet = writer.sheets['Vendor_Report']

            # Apply date format
            for cell in worksheet["G"][1:]:
                cell.number_format = 'DD/MM/YYYY'
            for cell in worksheet["I"][1:]:
                cell.number_format = 'DD/MM/YYYY'

            # Header Styling
            header_font = openpyxl.styles.Font(
                bold=True,
                color="FFFFFF"
            )

            header_fill = openpyxl.styles.PatternFill(
                start_color="4F81BD",
                end_color="4F81BD",
                fill_type="solid"
            )

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal="center"
                )

            # Alternate Row Styling & Column Width
            thin_border = openpyxl.styles.Side(
                border_style="thin",
                color="000000"
            )

            border = openpyxl.styles.Border(
                left=thin_border,
                right=thin_border,
                top=thin_border,
                bottom=thin_border
            )

            for row_idx, row in enumerate(
                worksheet.iter_rows(min_row=2, max_row=len(data)+1),
                start=2
            ):

                fill = openpyxl.styles.PatternFill(
                    start_color="DCE6F1",
                    end_color="DCE6F1",
                    fill_type="solid"
                ) if row_idx % 2 == 0 else None

                for cell in row:
                    if fill:
                        cell.fill = fill

                    cell.border = border

                # Auto width
                for idx, col in enumerate(df.columns):
                    try:
                        max_len = max(
                            df[col].astype(str).map(len).max(),
                            len(col)
                        ) + 4

                        worksheet.column_dimensions[
                            openpyxl.utils.get_column_letter(idx + 1)
                        ].width = min(max_len, 50)

                    except:
                        worksheet.column_dimensions[
                            openpyxl.utils.get_column_letter(idx + 1)
                        ].width = 20

                    worksheet.freeze_panes = 'A2'

            output.seek(0)
            filename = f"vendor_list_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
            print(f"INFO: [{datetime.now()}] Excel export completed: {filename}")
        
        # ✅ Safest alternative: Return raw bytes with explicit headers
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response

    except Exception as e:
        print(f"ERROR: [{datetime.now()}] Excel export failed: {str(e)}")
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()
#new Upaded Code on 15-05-2026
@super_admin_bp.route("/api/vendors/export/pdf", methods=["GET"])
def export_vendors_pdf():
    try:
        print(f"INFO: [{datetime.now()}] Initiating PDF export for Vendor Management.")
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                vendor_name,
                primary_contact,
                mail_id,
                address,
                status,
                created_by,
                DATE_FORMAT(created_at, '%d-%m-%Y') as created_on,
                modified_by,
                DATE_FORMAT(modified_on, '%d-%m-%Y') as modified_on
            FROM vendor_details 
            ORDER BY id DESC
        """)
        data = cursor.fetchall()

        if not data:
            return jsonify({"error": "No data available for export"}), 404

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Title Section
        title_style = styles['Title']
        title_style.fontSize = 20
        title_style.textColor = colors.HexColor("#1e40af")
        elements.append(Paragraph("Vendor Management Report", title_style))
        
        date_style = styles['Normal']
        date_style.alignment = 1
        elements.append(Paragraph(f"Exported on: {datetime.now().strftime('%B %d, %Y | %H:%M:%S')}", date_style))
        elements.append(Spacer(1, 24))

        # Table Construction
        table_data = [["Vendor Name", "Contact", "Status", "Created By", "Created On", "Modified By", "Modified On"]]
        
        for row in data:
            table_data.append([
                row['vendor_name'],
                row['primary_contact'],
                row['status'],
                row['created_by'] or 'System',
                row['created_on'] or '',
                row['modified_by'] or 'N/A',
                row['modified_on'] or 'N/A'
            ])

        # Table Formatting
        column_widths = [140, 90, 70, 100, 90, 100, 90]
        t = Table(table_data, colWidths=column_widths, repeatRows=1)
        
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(t)
        doc.build(elements)
        
        output.seek(0)
        filename = f"vendor_list_{datetime.now().strftime('%Y_%m_%d')}.pdf"
        print(f"INFO: [{datetime.now()}] PDF export completed: {filename}")
        
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')

    except Exception as e:
        print(f"ERROR: [{datetime.now()}] PDF export failed: {str(e)}")
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

        
@super_admin_bp.route("/api/vendors", methods=["GET"])
def get_vendors():

    organizer_id = request.args.get("organizer_id")
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if organizer_id:
        query = """
        SELECT id,vendor_name,primary_contact,mail_id,country,state,city,address,status,created_by,created_at as created_on,modified_by,modified_on
        FROM vendor_details
        WHERE organizer_id = %s
        ORDER BY id DESC
        """
        cursor.execute(query, (organizer_id,))
    else:
        query = """
        SELECT id,vendor_name,primary_contact,mail_id,country,state,city,address,status,created_by,created_at as created_on,modified_by,modified_on
        FROM vendor_details
        ORDER BY id DESC
        """
        cursor.execute(query)

    vendors = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(vendors)
@super_admin_bp.route("/api/vendor/<int:id>", methods=["GET"])
def view_vendor(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    print("Vendar",id)

    cursor.execute(
        "SELECT * FROM vendor_details WHERE id=%s",(id,)
    )

    vendor = cursor.fetchone()

    if vendor and vendor.get("bank_passbook"):
        vendor["bank_passbook"] = get_image_url(vendor["bank_passbook"])

    cursor.execute(
        "SELECT * FROM vendor_documents WHERE vendor_id=%s",(id,)
    )
    documents = cursor.fetchall()
    for d in documents:
        if d.get("document_file"):
            d["document_file"] = get_image_url(d["document_file"])

    vendor["documents"] = documents
    print("Vendar",vendor)

    cursor.close()
    conn.close()

    return jsonify(vendor)

@super_admin_bp.route("/api/delete_vendor/<int:id>", methods=["DELETE"])
def delete_vendor(id):
    try:
        print("Delete Vendor", id)
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Check if vendor exists
        cursor.execute("SELECT * FROM vendor_details WHERE id = %s", (id,))
        vendor = cursor.fetchone()

        if not vendor:
            return jsonify({
                "status": False,
                "message": "Vendor not found"
            }), 404

        # Delete vendor documents
        cursor.execute("DELETE FROM vendor_documents WHERE vendor_id = %s", (id,))
        
        # Delete vendor
        cursor.execute("DELETE FROM vendor_details WHERE id = %s", (id,))
        conn.commit()

        return jsonify({
            "status": True,
            "message": "Vendor deleted successfully"
        }), 200

    except Exception as e:
        print("ERROR:", str(e))
        return jsonify({
            "status": False,
            "message": "Failed to delete vendor",
            "error": str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@super_admin_bp.route("/api/venues_details", methods=["GET"])
def get_venues_detail():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        # Use 'Active' to match the database default
        cursor.execute("SELECT id, venue_name, city_name, address, state_name, country_name, pin_code FROM venues WHERE status='Active'")
        venues = cursor.fetchall()
        return jsonify(venues)
    except Exception as e:
        print("Error fetching venues details:", str(e))
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
@super_admin_bp.route("/api/create_policy", methods=["POST"])
def create_policy():
    try:
        data = request.json
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # ✅ Generate POLICY CODE (POL-1, POL-2...)
        cursor.execute("SELECT MAX(id) as max_id FROM policies")
        result = cursor.fetchone()

        new_number = (result["max_id"] or 0) + 1
        policy_code = f"POL-{new_number}"

        organizer_id = data.get("organizer_id")
        created_by = data.get("created_by", "System")
        
        if organizer_id and created_by == "System":
            cursor.execute("SELECT name FROM users WHERE id = %s", (organizer_id,))
            user_row = cursor.fetchone()
            if user_row:
                created_by = user_row["name"]

        # ✅ INSERT
        cursor.execute("""
            INSERT INTO policies 
            (policy_code, policy_name, policy_type, policy_group, description, organizer_id, status, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            policy_code,
            data.get("policy_name"),
            data.get("policy_type"),
            data.get("policy_group"),
            data.get("description"),
            organizer_id,
            data.get("status", "Active"),
            created_by
        ))

        conn.commit()

        return jsonify({
            "message": "Policy created successfully",
            "policy_code": policy_code
        }), 201

    except Exception as e:
        print("ERROR:", str(e))
        return jsonify({
            "error": "Failed to create policy",
            "details": str(e)
        }), 500

    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route("/api/update_policy/<int:id>", methods=["PUT"])
def update_policy(id):
    try:
        data = request.json
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        organizer_id = data.get("organizer_id")
        modified_by = data.get("modified_by", "System")
        
        if organizer_id and modified_by == "System":
            cursor.execute("SELECT name FROM users WHERE id = %s", (organizer_id,))
            user_row = cursor.fetchone()
            if user_row:
                modified_by = user_row["name"]

        cursor.execute("""
            UPDATE policies 
            SET policy_name=%s, policy_type=%s, policy_group=%s, description=%s, status=%s, modified_by=%s
            WHERE id=%s
        """, (
            data.get("policy_name"),
            data.get("policy_type"),
            data.get("policy_group"),
            data.get("description"),
            data.get("status", "Active"),
            modified_by,
            id
        ))

        conn.commit()

        return jsonify({"message": "Policy updated successfully"}), 200

    except Exception as e:
        print("Update Policy ERROR:", str(e))
        return jsonify({"error": "Failed to update policy", "details": str(e)}), 500

    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route("/api/delete_policy/<int:id>", methods=["DELETE"])
def delete_policy(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # ✅ Check if policy exists
        cursor.execute("SELECT * FROM policies WHERE id = %s", (id,))
        policy = cursor.fetchone()

        if not policy:
            return jsonify({
                "status": False,
                "message": "Policy not found"
            }), 404

        # ✅ Delete policy
        cursor.execute("DELETE FROM policies WHERE id = %s", (id,))
        conn.commit()

        return jsonify({
            "status": True,
            "message": "Policy deleted successfully"
        }), 200

    except Exception as e:
        print("ERROR:", str(e))
        return jsonify({
            "status": False,
            "message": "Failed to delete policy",
            "error": str(e)
        }), 500

    finally:
        cursor.close()
        conn.close()


# GET ALL
@super_admin_bp.route("/api/all-policies/<int:organizer_id>", methods=["GET"])
def get_policies(organizer_id):
    conn = None
    cursor = None

    try:
        print("Organizer",organizer_id)
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        query = "SELECT id, policy_code, policy_name, policy_type, policy_group, description, organizer_id, status, created_by, created_at as created_on, modified_by, modified_on FROM policies WHERE organizer_id = %s ORDER BY id DESC"
        cursor.execute(query, (organizer_id,))

        policies = cursor.fetchall()
        print(policies)

        return jsonify({
            "success": True,
            "data": policies
        }), 200

    except Exception as e:
        print("Error fetching policies:", str(e))
        return jsonify({
            "success": False,
            "message": "Failed to fetch policies",
            "error": str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# GET SINGLE
@super_admin_bp.route("/api/policy/<int:id>", methods=["GET"])
def get_policy(id):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM policies WHERE id=%s", (id,))
        policy = cursor.fetchone()
        return jsonify(policy)
    except Exception as e:
        print("Error fetching single policy:", str(e))
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
import datetime as dt

def format_date(date_str):
    if not date_str:
        return None

    formats = [
        "%Y-%m-%d",   # 2026-03-17
        "%d/%m/%Y",   # 17/03/2026
        "%d-%m-%Y",   # 17-03-2026
    ]

    for fmt in formats:
        try:
            return dt.datetime.strptime(date_str, fmt).date()
        except:
            continue

    print("Date format error: Unsupported format ->", date_str)
    return None

def format_time(time_str):
    if not time_str:
        return None

    formats = [
        "%H:%M",        # 14:30
        "%I:%M %p",     # 02:30 PM
    ]

    for fmt in formats:
        try:
            return dt.datetime.strptime(time_str, fmt).time()
        except:
            continue

    print("Time format error: Unsupported format ->", time_str)
    return None

def format_datetime(dt_str):
    try:
        if dt_str:
            return datetime.strptime(dt_str, "%Y-%m-%dT%H:%M")
    except Exception as e:
        print("Datetime error:", e)
    return None




import os

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
BASE_UPLOAD = os.path.abspath(os.path.join(BASE_DIR, "..", "uploads"))

PHOTO_FOLDER = os.path.join(BASE_UPLOAD, "photos")
VIDEO_FOLDER = os.path.join(BASE_UPLOAD, "videos")
DOC_FOLDER = os.path.join(BASE_UPLOAD, "documents")

os.makedirs(PHOTO_FOLDER, exist_ok=True)
os.makedirs(VIDEO_FOLDER, exist_ok=True)
os.makedirs(DOC_FOLDER, exist_ok=True)

def get_file_folder(file):
    content_type = file.content_type

    if content_type.startswith("image"):
        return PHOTO_FOLDER, "photos", "image"

    elif content_type.startswith("video"):
        return VIDEO_FOLDER, "videos", "video"

    else:
        return DOC_FOLDER, "documents", "document"
# =====================================================
# ✅ COMPLETE TRANSACTIONAL EVENT CREATION
# =====================================================
@super_admin_bp.route("/api/complete-event", methods=["POST"])
def complete_event_creation():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Access data from form fields (sent as JSON strings)
        event_details = json.loads(request.form.get("eventDetails", "{}"))
        booking_details = json.loads(request.form.get("booking", "{}"))
        layout_details = json.loads(request.form.get("layout", "{}"))
        terms_data = json.loads(request.form.get("terms", "[]"))
        vendors_data = json.loads(request.form.get("vendors", "{}"))
        food_provision = json.loads(request.form.get("foodProvision", '{"items": []}'))
        vehicle_provision = json.loads(request.form.get("vehicleProvision", '{"details": [], "addons": []}'))

        # 1. 🔹 INSERT EVENT DETAILS
        insert_event_query = """
        INSERT INTO event_details_table (
            category, event_name, description, amenities, tags,
            visibility, include_program,
            mail, whatsapp, print,
            visitor_mail, visitor_name, visitor_photo, visitor_mobile, document_proof,
            day_pass, is_international_include,
            aadhar, passport,
            welcome_kit, food, vehicle_pass, vehicle_number,
            event_type, occurrence,
            start_date, start_time, end_date, end_time,
            venue, address, created_by, user_id, status
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        event_values = (
            event_details.get("category"),
            event_details.get("eventName"),
            event_details.get("description"),
            event_details.get("amenities", []),
            event_details.get("tags", ""),
            event_details.get("visibility"),
            str(event_details.get("includeProgram", False)),
            event_details.get("mail", False),
            event_details.get("whatsapp", False),
            event_details.get("print", False),
            event_details.get("visitorMail", False),
            event_details.get("visitorName", False),
            event_details.get("visitorPhoto", False),
            event_details.get("visitorMobile", False),
            event_details.get("documentProof", False),
            event_details.get("dayPass", False),
            event_details.get("isInternationalInclude", False),
            event_details.get("aadhar", False),
            event_details.get("passport", False),
            event_details.get("welcomeKit", False),
            event_details.get("food", False),
            event_details.get("vehiclePass", False),
            event_details.get("vehicleNumber", False),
            event_type := event_details.get("eventType"),
            occurrence := event_details.get("occurrence"),
            format_date(event_details.get("startDate")),
            format_time(event_details.get("startTime")),
            format_date(event_details.get("endDate")),
            format_time(event_details.get("endTime")),
            event_details.get("venue"),
            event_details.get("address"),
            event_details.get("created_by"),
            event_details.get("user_id"),
            "PENDING" # Final state
        )
        cursor.execute(insert_event_query, event_values)
        event_id = cursor.lastrowid
        event_code = f"EVT-{str(event_id).zfill(4)}"
        cursor.execute("UPDATE event_details_table SET event_code=%s WHERE id=%s", (event_code, event_id))

        # 2. 🔹 INSERT BOOKING DETAILS
        insert_booking_query = """
        INSERT INTO event_booking_details (
            event_id, booking_start_date, booking_end_date, capacity, pass_type,
            title, title_type, title_selection, designation, designation_type, designation_selection,
            company, company_type, company_selection, entry_type, charge_type, max_pass,
            razorpay_key, include_tax, price_type, currency, early_bird_expire
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        booking_values = (
            event_id,
            format_date(booking_details.get("bookingStartDate")),
            format_date(booking_details.get("bookingEndDate")),
            booking_details.get("capacity"),
            booking_details.get("passType"),
            booking_details.get("title"), booking_details.get("titleType"), booking_details.get("titleSelection"),
            booking_details.get("designation"), booking_details.get("designationType"), booking_details.get("designationSelection"),
            booking_details.get("company"), booking_details.get("companyType"), booking_details.get("companySelection"),
            booking_details.get("entryType"),
            booking_details.get("chargeType"),
            booking_details.get("maxPass"),
            booking_details.get("razorpayKey"),
            booking_details.get("includeTax", False),
            booking_details.get("priceType"),
            booking_details.get("currency"),
            format_datetime(booking_details.get("earlyBirdExpire"))
        )
        cursor.execute(insert_booking_query, booking_values)

        # 3. 🔹 INSERT LAYOUT MASTER
        cursor.execute("""
            INSERT INTO event_layout (
                event_id, floor_type, day_based, person_pass, include_tax
            ) VALUES (%s, %s, %s, %s, %s)
        """, (
            event_id,
            layout_details.get("floorType"),
            layout_details.get("dayBased"),
            layout_details.get("personPass"),
            layout_details.get("includeTax")
        ))

        # 4. 🔹 INSERT STALLS
        stalls = layout_details.get("stalls", [])
        for stall in stalls:
            cursor.execute("""
                INSERT INTO event_stalls (
                    event_id, stall_name, stall_size, size_range, visibility,
                    stall_type, price_inr, price_usd, prime_seat, prime_price_inr, prime_price_usd
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                event_id,
                stall.get("stallName"),
                stall.get("size"),
                stall.get("sizeRange"),
                stall.get("visibility"),
                stall.get("type"),
                stall.get("priceINR"),
                stall.get("priceUSD"),
                stall.get("primeSeat"),
                stall.get("primePriceINR"),
                stall.get("primePriceUSD")
            ))

        # 5. 🔹 INSERT STALL AMENITIES
        amenities = layout_details.get("amenities", [])
        for a in amenities:
            cursor.execute("""
                INSERT INTO stall_amenities (
                    event_id, stall_name, amenity, qty
                ) VALUES (%s, %s, %s, %s)
            """, (
                event_id,
                a.get("stallName"),
                a.get("amenity"),
                a.get("qty")
            ))

        # 6. 🔹 HANDLE FILES (BANNER & DOCUMENTS)
        banner = request.files.get("banner")
        if banner:
            folder, relative_cat, _ = get_file_folder(banner)
            filename = secure_filename(banner.filename)
            save_path = os.path.join(folder, filename)
            db_path = f"/uploads/{relative_cat}/{filename}"
            banner.save(save_path)
            cursor.execute("INSERT INTO event_files (event_id, file_name, file_path, file_type) VALUES (%s, %s, %s, %s)", (event_id, filename, db_path, "banner"))

        doc_count = int(request.form.get("doc_count", 0))
        for i in range(doc_count):
            file = request.files.get(f"docs_{i}")
            if file:
                folder, relative_cat, file_category = get_file_folder(file)
                filename = secure_filename(file.filename)
                save_path = os.path.join(folder, filename)
                db_path = f"/uploads/{relative_cat}/{filename}"
                file.save(save_path)
                cursor.execute("""
                    INSERT INTO event_files (event_id, file_name, file_path, file_type, doc_type, doc_number)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (event_id, filename, db_path, file_category, request.form.get(f"doc_type_{i}"), request.form.get(f"doc_number_{i}")))

        # 7. 🔹 INSERT TERMS
        for term in terms_data:
            cursor.execute("""
                INSERT INTO event_terms (event_id, policy_group, policy_type, policy_name, is_default)
                VALUES (%s, %s, %s, %s, %s)
            """, (event_id, term.get("policyGroup"), term.get("policyType"), term.get("policyName"), term.get("isDefault", False)))

        # 8. 🔹 INSERT VENDORS, SPONSORS, GUESTS
        for v in vendors_data.get("vendors", []):
            cursor.execute("INSERT INTO event_vendors (event_id, vendor_type, vendor_name, pass_count) VALUES (%s, %s, %s, %s)", (event_id, v.get("vendorType"), v.get("vendorName"), v.get("passCount", 0)))
        for s in vendors_data.get("sponsors", []):
            cursor.execute("INSERT INTO event_sponsors (event_id, sponsor_name, sponsorship_type) VALUES (%s, %s, %s)", (event_id, s.get("sponsorName"), s.get("sponsorship")))
        for g in vendors_data.get("guests", []):
            cursor.execute("INSERT INTO event_guests (event_id, guest_name, designation, contact, image) VALUES (%s, %s, %s, %s, %s)", (event_id, g.get("name"), g.get("designation"), g.get("contact"), g.get("image")))

        # 9. 🔹 INSERT FOOD ITEMS
        for item in food_provision.get("items", []):
            cursor.execute("""
                INSERT INTO event_food_items (event_id, caterer_name, meal_type, food_type, price_inr, price_usd, menu_details)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (event_id, item.get("catererName"), item.get("mealType"), item.get("foodType"), item.get("priceINR") or 0, item.get("priceUSD") or 0, item.get("menuDetails")))

        # 10. 🔹 INSERT VEHICLE DETAILS
        for item in vehicle_provision.get("details", []):
            cursor.execute("""
                INSERT INTO event_vehicle_details (event_id, vehicle_type, price_inr, price_usd)
                VALUES (%s, %s, %s, %s)
            """, (event_id, item.get("vehicleType"), item.get("priceINR") or 0, item.get("priceUSD") or 0))

        # 11. 🔹 INSERT VEHICLE ADDONS
        for item in vehicle_provision.get("addons", []):
            cursor.execute("""
                INSERT INTO event_vehicle_addons (event_id, is_parent, addon_name, price)
                VALUES (%s, %s, %s, %s)
            """, (event_id, item.get("isParent", False), item.get("addOnName"), item.get("price") or 0))

        conn.commit()
        return jsonify({"message": "Event created and submitted successfully", "event_id": event_id}), 201

    except Exception as e:
        conn.rollback()
        print("TRANSACTION ERROR:", str(e))
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()




@super_admin_bp.route("/api/all-policies", methods=["GET"])
def get_all_policies():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT policy_group, policy_type, policy_name
            FROM policies
            WHERE status='Active'
        """)

        rows = cursor.fetchall()

        # 🔥 Convert to nested JSON
        data = {}

        for row in rows:
            group = row["policy_group"]
            type_ = row["policy_type"]
            name = row["policy_name"]

            if group not in data:
                data[group] = {}

            if type_ not in data[group]:
                data[group][type_] = []

            data[group][type_].append(name)

        return jsonify(data)

    except Exception as e:
        print("Error in get_all_policies:", str(e))
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
# ✅ Get Vendor Types
@super_admin_bp.route("/api/get-vendor-types", methods=["GET"])
def get_vendor_types():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT id,vendor_type FROM vendor_details")
        data = cursor.fetchall()

        return jsonify(data)
    except Exception as e:
        print("Error in get_vendor_types:", str(e))
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# ✅ Get Vendor Names based on type
@super_admin_bp.route("/api/get-vendor-names/<vendor_type>", methods=["GET"])
def get_vendor_names(vendor_type):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(
        "SELECT vendor_name FROM vendor_details WHERE vendor_type=%s",
        (vendor_type,)
    )
    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(data)


# ✅ Get Sponsor Names
@super_admin_bp.route("/api/get-sponsor-names", methods=["GET"])
def get_sponsor_names():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT DISTINCT sponsor_name FROM sponsors_details")
    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(data)
# =========================
# SAVE VENDORS + SPONSORS + GUESTS
# =========================
@super_admin_bp.route("/api/save-vendors-sponsors", methods=["POST"])
def save_vendors():
    try:
        data = request.json
        event_id = data.get("event_id")

        vendors = data.get("vendors", [])
        sponsors = data.get("sponsors", [])
        guests = data.get("guests", [])

        conn = get_db_connection()
        cursor = conn.cursor()

        # =========================
        # CLEAR OLD DATA (IMPORTANT)
        # =========================
        cursor.execute("DELETE FROM event_vendors WHERE event_id=%s", (event_id,))
        cursor.execute("DELETE FROM event_sponsors WHERE event_id=%s", (event_id,))
        cursor.execute("DELETE FROM event_guests WHERE event_id=%s", (event_id,))

        # =========================
        # INSERT VENDORS
        # =========================
        for v in vendors:
            cursor.execute("""
                INSERT INTO event_vendors (event_id, vendor_type, vendor_name, pass_count)
                VALUES (%s, %s, %s, %s)
            """, (
                event_id,
                v.get("vendorType"),
                v.get("vendorName"),
                v.get("passCount", 0)
            ))

        # =========================
        # INSERT SPONSORS
        # =========================
        for s in sponsors:
            cursor.execute("""
                INSERT INTO event_sponsors (event_id, sponsor_name, sponsorship_type)
                VALUES (%s, %s, %s)
            """, (
                event_id,
                s.get("sponsorName"),
                s.get("sponsorship")
            ))

        # =========================
        # INSERT GUESTS
        # =========================
        for g in guests:
            cursor.execute("""
                INSERT INTO event_guests (event_id, guest_name, designation, contact, image)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                event_id,
                g.get("name"),
                g.get("designation"),
                g.get("contact"),
                g.get("image")
            ))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"message": "Vendors, Sponsors, Guests saved successfully"}), 200

    except Exception as e:
        print("ERROR:", str(e))
        return jsonify({"error": str(e)}), 500

import datetime as dt

@super_admin_bp.route("/event/final-submit", methods=["POST"])
def final_submit():
    try:
        data = request.json
        event_id = data.get("event_id")

        print("DATA:", data)

        if not event_id:
            return jsonify({"error": "event_id is required"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE event_details_table
            SET status = %s
            WHERE id = %s
        """, ("PENDING", event_id))

        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({
            "message": "Event submitted successfully",
            "status": "PENDING",
            "event_id": event_id
        }), 200

    except Exception as e:
        print("FINAL SUBMIT ERROR:", str(e))
        return jsonify({"error": str(e)}), 500

from flask import send_from_directory
# =====================================================
# ✅ ABSOLUTE UPLOAD FOLDER
# =====================================================
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.abspath(os.path.join(BASE_DIR, "..", "uploads"))

# Example:
# D:/Project/app/super_admin → BASE_DIR
# D:/Project/app/uploads → UPLOAD_FOLDER


# =========================
# [CHECK] URL HELPER
# =========================
def get_image_url(path):
    if not path:
        return None
    file_path = str(path).replace("\\", "/")
    
    # If it contains "uploads/", strip everything before it to make it relative
    if "uploads/" in file_path:
        relative_part = file_path.split("uploads/")[-1]
        return f"https://eventsapi.sportalytics.in/uploads/{relative_part.lstrip('/')}"
    
    # Fallback for paths that don't have uploads/
    # If it starts with http or data URI, return it unmodified
    if file_path.startswith(("http://", "https://", "data:")):
        return file_path
    
    return f"https://eventsapi.sportalytics.in/{file_path.lstrip('/')}"


# =====================================================
# ✅ SERVE UPLOADED FILES (IMAGES / VIDEOS / DOCS)
# =====================================================
@super_admin_bp.route("/uploads/<path:filename>")
def serve_file(filename):
    try:
        return send_from_directory(UPLOAD_FOLDER, filename)
    except Exception as e:
        print("FILE ERROR:", e)
        return jsonify({"error": "File not found"}), 404

# =====================================================
# ✅ GET EVENTS WITH BANNER IMAGE
# =====================================================
@super_admin_bp.route("/get-events", methods=["GET"])
def get_events():
    try:
        organizer = request.args.get('organizer')
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        if organizer:
            query = """
            SELECT 
                e.id, e.status, e.event_name, e.category, e.description, e.amenities, e.tags,
                e.visibility, e.include_program, e.mail, e.whatsapp, e.print,
                e.visitor_mail, e.visitor_name, e.visitor_photo, e.visitor_mobile, e.document_proof,
                e.day_pass, e.is_international_include, e.aadhar, e.passport,
                e.welcome_kit, e.food, e.event_type, e.occurrence,
                e.start_date, e.start_time, e.end_date, e.end_time,
                e.venue, e.address, e.created_by,
                f.file_path, f.file_type AS banner_type,
                b.booking_start_date, b.booking_end_date, b.capacity, b.pass_type, b.entry_type, b.charge_type
            FROM event_details_table e
            LEFT JOIN event_files f 
                ON e.id = f.event_id AND f.file_type = 'banner'
            LEFT JOIN event_booking_details b
                ON e.id = b.event_id
            WHERE e.user_id = %s
            ORDER BY e.id DESC
            """
            cursor.execute(query, (organizer,))
        else:
            query = """
            SELECT 
                e.id, e.status, e.event_name, e.category, e.description, e.amenities, e.tags,
                e.visibility, e.include_program, e.mail, e.whatsapp, e.print,
                e.visitor_mail, e.visitor_name, e.visitor_photo, e.visitor_mobile, e.document_proof,
                e.day_pass, e.is_international_include, e.aadhar, e.passport,
                e.welcome_kit, e.food, e.event_type, e.occurrence,
                e.start_date, e.start_time, e.end_date, e.end_time,
                e.venue, e.address, e.created_by,
                f.file_path, f.file_type AS banner_type,
                b.booking_start_date, b.booking_end_date, b.capacity, b.pass_type, b.entry_type, b.charge_type
            FROM event_details_table e
            LEFT JOIN event_files f 
                ON e.id = f.event_id AND f.file_type = 'banner'
            LEFT JOIN event_booking_details b
                ON e.id = b.event_id
            ORDER BY e.id DESC
            """
            cursor.execute(query)
            
        data = cursor.fetchall()
        
        events_dict = {}

        for row in data:

            # =========================
            # ✅ DATE & TIME FIXES
            # =========================
            fields_to_fix = ["start_date", "end_date", "booking_start_date", "booking_end_date"]
            for f in fields_to_fix:
                if row.get(f):
                    row[f] = row[f].strftime("%Y-%m-%d")

            for tf in ["start_time", "end_time"]:
                if row.get(tf):
                    if isinstance(row[tf], timedelta):
                        total_seconds = int(row[tf].total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        row[tf] = f"{hours:02}:{minutes:02}"
                    else:
                        row[tf] = str(row[tf])

            # =========================
            # [CHECK] IMAGE PATH FIX (IMPORTANT)
            # =========================
            banner_url = get_image_url(row.get("file_path"))
                
            event_id = row["id"]
            if event_id not in events_dict:
                row["banner_url"] = banner_url
                row["images"] = []
                if banner_url:
                    row["images"].append({"url": banner_url, "banner_type": row["banner_type"]})
                events_dict[event_id] = row
            else:
                if banner_url:
                    if not any(img["url"] == banner_url for img in events_dict[event_id]["images"]):
                        events_dict[event_id]["images"].append({"url": banner_url, "banner_type": row["banner_type"]})

        final_data = list(events_dict.values())

        cursor.close()
        conn.close()

        return jsonify(final_data), 200

    except Exception as e:
        print("ERROR:", e)
        return jsonify({"error": str(e)}), 500
@super_admin_bp.route("/home/get-events", methods=["GET"])
def get_home_events():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        query = """
        SELECT 
            e.id,
            e.status,
            e.event_name,
            e.category,
            e.start_date,
            e.start_time,
            e.venue,
            e.address,
            f.file_path,
            f.file_type AS banner_type,
            b.capacity,
            b.currency,
            b.entry_type
        FROM event_details_table e
        LEFT JOIN event_files f 
            ON e.id = f.event_id AND f.file_type = 'banner'
        LEFT JOIN event_booking_details b
            ON e.id = b.event_id
        WHERE e.status = 'APPROVED' 
        """
        #WHERE e.status = 'APPROVED' AND e.start_date >= CURDATE()

        cursor.execute(query)
        data = cursor.fetchall()

        events_dict = {}

        for row in data:

            # =========================
            # ✅ DATE FIX
            # =========================
            if row["start_date"]:
                row["start_date"] = row["start_date"].strftime("%Y-%m-%d")

            # =========================
            # ✅ TIME FIX
            # =========================
            if row["start_time"]:
                if isinstance(row["start_time"], timedelta):
                    total_seconds = int(row["start_time"].total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    row["start_time"] = f"{hours:02}:{minutes:02}"
                else:
                    row["start_time"] = str(row["start_time"])

            # =========================
            # [CHECK] IMAGE PATH FIX (IMPORTANT)
            # =========================
            banner_url = get_image_url(row.get("file_path"))
                
            event_id = row["id"]
            if event_id not in events_dict:
                row["banner_url"] = banner_url
                row["images"] = []
                if banner_url:
                    row["images"].append({"url": banner_url, "banner_type": row["banner_type"]})
                events_dict[event_id] = row
            else:
                if banner_url:
                    if not any(img["url"] == banner_url for img in events_dict[event_id]["images"]):
                        events_dict[event_id]["images"].append({"url": banner_url, "banner_type": row["banner_type"]})

        final_data = list(events_dict.values())

        cursor.close()
        conn.close()

        return jsonify(final_data), 200


    except Exception as e:
        print("ERROR:", e)
        return jsonify({"error": str(e)}), 500


# --- timedelta handled at top ---

@super_admin_bp.route('/booking/event/<int:event_id>', methods=['GET'])
def get_event_booking(event_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM event_details_table WHERE id=%s", (event_id,))
    event = cursor.fetchone()

    if event:
        for key, value in event.items():
            if isinstance(value, timedelta):
                # Convert to HH:MM:SS format
                event[key] = str(value)

    return jsonify(event)

# ─── Get Full Event Details for Editing ───────────────────────────────────────
@super_admin_bp.route("/api/event-full-details/<int:event_id>", methods=["GET"])
def get_event_full_details(event_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # 1. Core Details
        cursor.execute("SELECT * FROM event_details_table WHERE id = %s", (event_id,))
        details = cursor.fetchone()
        if not details:
            return jsonify({"error": "Event not found"}), 404

        # Date/Time formatting
        for key, val in details.items():
            if isinstance(val, (date, datetime)):
                details[key] = val.strftime("%Y-%m-%d")
            elif isinstance(val, timedelta):
                total_seconds = int(val.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                details[key] = f"{hours:02}:{minutes:02}"

        # 2. Booking Details
        cursor.execute("SELECT * FROM event_booking_details WHERE event_id = %s", (event_id,))
        booking = cursor.fetchone()
        if booking:
            for key, val in booking.items():
                if isinstance(val, datetime) and key == "early_bird_expire":
                    booking[key] = val.strftime("%Y-%m-%dT%H:%M")
                elif isinstance(val, (date, datetime)):
                    booking[key] = val.strftime("%Y-%m-%d")

        # 3. Layout Details
        cursor.execute("SELECT * FROM event_layout WHERE event_id = %s", (event_id,))
        layout_master = cursor.fetchone()
        
        cursor.execute("SELECT * FROM event_stalls WHERE event_id = %s", (event_id,))
        stalls = cursor.fetchall()
        
        cursor.execute("SELECT * FROM stall_amenities WHERE event_id = %s", (event_id,))
        amenities = cursor.fetchall()

        # 4. Files
        cursor.execute("SELECT * FROM event_files WHERE event_id = %s", (event_id,))
        files = cursor.fetchall()
        
        for f in files:
            f["url"] = get_image_url(f.get("file_path"))

        # 5. Terms
        cursor.execute("SELECT * FROM event_terms WHERE event_id = %s", (event_id,))
        terms = cursor.fetchall()

        # 6. Vendors, Sponsors, Guests
        cursor.execute("SELECT * FROM event_vendors WHERE event_id = %s", (event_id,))
        vendors = cursor.fetchall()
        
        cursor.execute("SELECT * FROM event_sponsors WHERE event_id = %s", (event_id,))
        sponsors = cursor.fetchall()
        
        cursor.execute("SELECT * FROM event_guests WHERE event_id = %s", (event_id,))
        guests = cursor.fetchall()
        for g in guests:
            g["image_url"] = get_image_url(g.get("image"))
        
        # 7. Food Items & Vehicle Details
        cursor.execute("SELECT * FROM event_food_items WHERE event_id = %s", (event_id,))
        food_items = cursor.fetchall()
        
        cursor.execute("SELECT * FROM event_vehicle_details WHERE event_id = %s", (event_id,))
        vehicle_details = cursor.fetchall()

        cursor.execute("SELECT * FROM event_vehicle_addons WHERE event_id = %s", (event_id,))
        vehicle_addons = cursor.fetchall()

        result = {
            "details": details,
            "booking": booking,
            "layout": {
                "master": layout_master,
                "stalls": stalls,
                "amenities": amenities
            },
            "files": files,
            "terms": terms,
            "vendor_data": {
                "vendors": vendors,
                "sponsors": sponsors,
                "guests": guests
            },
            "food_items": food_items,
            "vehicle_details": vehicle_details,
            "vehicle_addons": vehicle_addons
        }

        return jsonify(result)

    except Exception as e:
        print("GET FULL DETAILS ERROR:", str(e))
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route('/api/admin/booking/<int:id>', methods=['GET'])
def get_single_booking(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        query = """
            SELECT * FROM Exhibitor_stall_bookings
            WHERE id = %s
        """
        cursor.execute(query, (id,))
        booking = cursor.fetchone()

        if not booking:
            return jsonify({"message": "Booking not found"}), 404

        booking["visiting_card_url"] = get_image_url(booking.get("visiting_card"))

        return jsonify(booking), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@super_admin_bp.route('/api/admin/bookings', methods=['GET'])
def get_all_bookings():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM Exhibitor_stall_bookings ORDER BY created_at DESC")
        bookings = cursor.fetchall()

        for b in bookings:
            b["visiting_card_url"] = get_image_url(b.get("visiting_card"))

        return jsonify(bookings), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
#new Add the Function 
@super_admin_bp.route('/api/admin/bookings/event/<int:event_id>', methods=['GET'])
def get_bookings_by_event(event_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        print("Event id", event_id)

        cursor.execute("SELECT * FROM Exhibitor_stall_bookings WHERE event_id = %s ORDER BY created_at DESC", (event_id,))
        bookings = cursor.fetchall()

        for b in bookings:
            b["visiting_card_url"] = get_image_url(b.get("visiting_card"))

        return jsonify(bookings), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@super_admin_bp.route('/api/admin/update-booking-status/<int:id>', methods=['PUT'])
def update_booking_status(id):
    try:
        data = request.json
        status = data.get("status")

        conn = get_db_connection()
        cursor = conn.cursor()

        query = """
            UPDATE Exhibitor_stall_bookings
            SET status = %s
            WHERE id = %s
        """
        cursor.execute(query, (status, id))
        conn.commit()

        return jsonify({"message": "Status updated successfully"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@super_admin_bp.route('/api/delete-event/<int:id>', methods=['DELETE'])
def delete_event(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if event exists
        cursor.execute("SELECT id FROM event_details_table WHERE id = %s", (id,))
        if not cursor.fetchone():
            return jsonify({"status": False, "message": "Event not found"}), 404

        cursor.execute("DELETE FROM event_details_table WHERE id = %s", (id,))
        
        # Optionally we might want to delete from event_files, event_booking_details, etc., if cascading is not set up
        # but deleting from event_details_table is the primary action.
        cursor.execute("DELETE FROM event_booking_details WHERE event_id = %s", (id,))
        cursor.execute("DELETE FROM event_vendors WHERE event_id = %s", (id,))
        cursor.execute("DELETE FROM event_sponsors WHERE event_id = %s", (id,))
        cursor.execute("DELETE FROM event_guests WHERE event_id = %s", (id,))
        cursor.execute("DELETE FROM event_terms WHERE event_id = %s", (id,))
        cursor.execute("DELETE FROM event_files WHERE event_id = %s", (id,))
        cursor.execute("DELETE FROM event_layout WHERE event_id = %s", (id,))
        cursor.execute("DELETE FROM event_stalls WHERE event_id = %s", (id,))
        
        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({"status": True, "message": "Event deleted successfully"}), 200

    except Exception as e:
        print("DELETE EVENT ERROR:", str(e))
        return jsonify({"status": False, "message": str(e)}), 500



 
# GET TASKS
@super_admin_bp.route('/api/get-tasks', methods=['GET'])
def get_tasks():
    try:
        conn = get_db_connection()   # ✅ FIXED
        cursor = conn.cursor(dictionary=True)
 
        cursor.execute("SELECT * FROM todo_tasks ORDER BY created_at DESC")
        tasks = cursor.fetchall()
 
        for task in tasks:
            if isinstance(task['start_date'], date):
                task['start_date'] = task['start_date'].strftime('%Y-%m-%d')
            if isinstance(task['end_date'], date):
                task['end_date'] = task['end_date'].strftime('%Y-%m-%d')
            if task.get('created_at'):
                task['created_at'] = str(task['created_at'])
 
        cursor.close()
        conn.close()
 
        return jsonify({'success': True, 'data': tasks})
 
    except Exception as e:
        print(f"Error in get_tasks: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

 
 
# CREATE TASK
@super_admin_bp.route('/api/create-tasks', methods=['POST'])
def create_tasks():
    try:
        data = request.get_json()
 
        task_name = data.get('task_name', '').strip()
        task_description = data.get('task_description', '').strip()
        todo_items = data.get('todo_items', [])
 
        if not task_name or not todo_items:
            return jsonify({'success': False, 'error': 'task_name and todo_items are required'}), 400
 
        conn = get_db_connection()   # ✅ FIXED
        cursor = conn.cursor()
 
        inserted = []
 
        for item in todo_items:
            cursor.execute("""
                INSERT INTO todo_tasks
                (task_name, task_description, todo_list_name, start_date, end_date,
                 assigned_to, status, complete_percent, remarks)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                task_name,
                task_description,
                item.get('todo_list_name'),
                item.get('start_date') or None,
                item.get('end_date') or None,
                item.get('assigned_to'),
                item.get('status', 'In-Progress'),
                int(item.get('complete_percent') or 0),
                item.get('remarks', '')
            ))
            inserted.append(cursor.lastrowid)

 
        conn.commit()
        cursor.close()
        conn.close()
 
        return jsonify({
            'success': True,
            'inserted_ids': inserted,
            'count': len(inserted)
        }), 201
 
    except Exception as e:
        print(f"Error in create_tasks: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500



@super_admin_bp.route("/api/add-on-spot-events", methods=["GET"])
def get_all_active_events():

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT 
                event_code,
                event_name,
                start_date,
                end_date
            FROM event_details_table
            WHERE status = 'APPROVED'
        """)

        data = cursor.fetchall()

        return jsonify(data)

    except Exception as e:
        return jsonify({"error": str(e)})

    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route("/api/events-check-in", methods=["GET"])
def get_events_checkin():

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT 
                id,
                event_code,
                event_name,
                start_date,
                end_date
            FROM event_details_table
            WHERE status = 'APPROVED'
        """)

        data = cursor.fetchall()

        return jsonify(data)

    except Exception as e:
        return jsonify({"error": str(e)})

    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route("/api/program-verification/events", methods=["GET"])
def get_program_verification_events():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT 
                id,
                event_code,
                event_name
            FROM event_details_table
            WHERE status = 'APPROVED'
            ORDER BY created_at DESC
        """)

        data = cursor.fetchall()

        return jsonify(data)

    except Exception as e:
        return jsonify({"error": str(e)})

    finally:
        cursor.close()
        conn.close()

# ─── Page 1: List all approved events ────────────────────────────────────────
@super_admin_bp.route("/api/message-greetings", methods=["GET"])
def get_all_message_greetings():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT
                id,
                event_name,
                start_date,
                end_date
            FROM event_details_table
            WHERE status = 'APPROVED'
            ORDER BY start_date DESC
        """)
        data = cursor.fetchall()
        # Serialize dates to strings
        for row in data:
            if isinstance(row.get("start_date"), date):
                row["start_date"] = row["start_date"].strftime("%d/%m/%Y")
            if isinstance(row.get("end_date"), date):
                row["end_date"] = row["end_date"].strftime("%d/%m/%Y")
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()
 
 
# ─── Page 2: Get messages for a specific event ────────────────────────────────
@super_admin_bp.route("/api/message-greetings/<int:event_id>/messages", methods=["GET"])
def get_messages_by_event(event_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT
                id,
                event_id,
                message_group,
                topics,
                sub_topics,
                description,
                image_path,
                type,
                created_at
            FROM messages_greetings_table
            WHERE event_id = %s
            ORDER BY created_at DESC
        """, (event_id,))
        data = cursor.fetchall()
        for row in data:
            if row.get("created_at"):
                row["created_at"] = row["created_at"].strftime("%d/%m/%Y %H:%M")
            row["image_url"] = get_image_url(row.get("image_path"))
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()
 
 
# ─── Page 2: Save a new message ──────────────────────────────────────────────
@super_admin_bp.route("/api/message-greetings/<int:event_id>/messages", methods=["POST"])
def save_message(event_id):
    try:
        data = request.get_json()
        message_group = data.get("message_group")
        topics = data.get("topics", "")
        sub_topics = data.get("sub_topics", "")
        description = data.get("description", "")
        image_path = data.get("image_path", "")
        msg_type = data.get("type", "Messages")  # 'Messages' or 'Greetings'
 
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO messages_greetings_table
                (event_id, message_group, topics, sub_topics, description, image_path, type)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (event_id, message_group, topics, sub_topics, description, image_path, msg_type))
        conn.commit()
        new_id = cursor.lastrowid
        return jsonify({"success": True, "id": new_id}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()
 
 
# ─── Page 2: Delete a message ────────────────────────────────────────────────
@super_admin_bp.route("/api/message-greetings/messages/<int:message_id>", methods=["DELETE"])
def delete_message(message_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM messages_greetings_table WHERE id = %s", (message_id,))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# ─── Page 2: Update a message ────────────────────────────────────────────────
@super_admin_bp.route("/api/message-greetings/messages/<int:message_id>", methods=["PUT"])
def update_message(message_id):
    try:
        data = request.get_json()
        message_group = data.get("message_group")
        topics = data.get("topics", "")
        sub_topics = data.get("sub_topics", "")
        description = data.get("description", "")
        image_path = data.get("image_path", "")
        msg_type = data.get("type", "Messages")

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE messages_greetings_table
            SET message_group=%s, topics=%s, sub_topics=%s, description=%s, image_path=%s, type=%s
            WHERE id=%s
        """, (message_group, topics, sub_topics, description, image_path, msg_type, message_id))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# ─── Update Event ────────────────────────────────────────────────────────────
@super_admin_bp.route("/api/update-event/<int:event_id>", methods=["PUT"])
def update_event(event_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        if 'eventDetails' in request.form:
             import json
             event_details = json.loads(request.form.get('eventDetails'))
             booking_details = json.loads(request.form.get('booking', '{}'))
             layout_details = json.loads(request.form.get('layout', '{}'))
             terms_data = json.loads(request.form.get('terms', '[]'))
             vendors_data = json.loads(request.form.get('vendors', '{}'))
             food_provision = json.loads(request.form.get('foodProvision', '{"items": []}'))
             vehicle_provision = json.loads(request.form.get('vehicleProvision', '{"details": [], "addons": []}'))

             # 1. Update Core Details
             cursor.execute("""
                UPDATE event_details_table 
                SET event_name=%s, category=%s, description=%s, amenities=%s, tags=%s,
                    include_program=%s, visibility=%s, mail=%s, whatsapp=%s, print=%s,
                    visitor_mail=%s, visitor_name=%s, visitor_photo=%s, visitor_mobile=%s, document_proof=%s,
                    day_pass=%s, is_international_include=%s, aadhar=%s, passport=%s,
                    welcome_kit=%s, food=%s, vehicle_pass=%s, vehicle_number=%s, event_type=%s, occurrence=%s,
                    start_date=%s, start_time=%s, end_date=%s, end_time=%s, venue=%s, address=%s
                WHERE id=%s
             """, (
                 event_details.get('eventName'),
                 event_details.get('category'),
                 event_details.get('description'),
                 event_details.get('amenities'),
                 event_details.get('tags'),
                 str(event_details.get('includeProgram')),
                 event_details.get('visibility'),
                 event_details.get('mail'), event_details.get('whatsapp'), event_details.get('print'),
                 event_details.get('visitorMail'), event_details.get('visitorName'), event_details.get('visitorPhoto'), event_details.get('visitorMobile'), event_details.get('documentProof'),
                 event_details.get('dayPass'), event_details.get('isInternationalInclude'), event_details.get('aadhar'), event_details.get('passport'),
                 event_details.get('welcomeKit'), event_details.get('food'),
                 event_details.get('vehiclePass'), event_details.get('vehicleNumber', False),
                 event_details.get('eventType'), event_details.get('occurrence'),
                 format_date(event_details.get('startDate')), format_time(event_details.get('startTime')),
                 format_date(event_details.get('endDate')), format_time(event_details.get('endTime')),
                 event_details.get('venue'), event_details.get('address'),
                 event_id
             ))

             # 2. Update Booking
             if booking_details:
                   cursor.execute("""
                     UPDATE event_booking_details
                     SET booking_start_date=%s, booking_end_date=%s, capacity=%s, pass_type=%s, entry_type=%s, charge_type=%s,
                         max_pass=%s, razorpay_key=%s, include_tax=%s, price_type=%s, currency=%s, early_bird_expire=%s
                     WHERE event_id=%s
                   """, (
                       format_date(booking_details.get('bookingStartDate')),
                       format_date(booking_details.get('bookingEndDate')),
                       booking_details.get('capacity'),
                       booking_details.get('passType'),
                       booking_details.get('entryType'),
                       booking_details.get('chargeType'),
                       booking_details.get('maxPass'),
                       booking_details.get('razorpayKey'),
                       booking_details.get('includeTax', False),
                       booking_details.get('priceType'),
                       booking_details.get('currency'),
                       format_datetime(booking_details.get('earlyBirdExpire')),
                       event_id
                   ))

             # 3. Update Layout Master & Stalls (Delete and Re-insert)
             if layout_details:
                 # Check if layout master exists
                 cursor.execute("SELECT COUNT(*) FROM event_layout WHERE event_id = %s", (event_id,))
                 layout_exists = cursor.fetchone()[0] > 0
                 
                 day_based = 1 if layout_details.get("dayBased") in [True, 1, "1", "true", "True"] else 0
                 include_tax = 1 if layout_details.get("includeTax") in [True, 1, "1", "true", "True"] else 0
                 
                 person_pass = layout_details.get("personPass")
                 if person_pass == "" or person_pass is None:
                     person_pass = 0
                 else:
                     try:
                         person_pass = int(person_pass)
                     except ValueError:
                         person_pass = 0

                 if layout_exists:
                     cursor.execute("""
                        UPDATE event_layout
                        SET floor_type=%s, day_based=%s, person_pass=%s, include_tax=%s
                        WHERE event_id=%s
                     """, (
                         layout_details.get("floorType"),
                         day_based,
                         person_pass,
                         include_tax,
                         event_id
                     ))
                 else:
                     cursor.execute("""
                        INSERT INTO event_layout (event_id, floor_type, day_based, person_pass, include_tax)
                        VALUES (%s, %s, %s, %s, %s)
                     """, (
                         event_id,
                         layout_details.get("floorType"),
                         day_based,
                         person_pass,
                         include_tax
                     ))

                 cursor.execute("DELETE FROM stall_amenities WHERE event_id = %s", (event_id,))
                 cursor.execute("DELETE FROM event_stalls WHERE event_id = %s", (event_id,))
                 
                 for stall in layout_details.get('stalls', []):
                     cursor.execute("""
                        INSERT INTO event_stalls 
                        (event_id, stall_name, stall_size, size_range, visibility, stall_type, price_inr, price_usd, prime_seat, prime_price_inr, prime_price_usd)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                     """, (
                         event_id, 
                         stall.get('stallName'), 
                         stall.get('size'), 
                         stall.get('sizeRange'), 
                         stall.get('visibility'), 
                         stall.get('type'), 
                         stall.get('priceINR'), 
                         stall.get('priceUSD'), 
                         1 if stall.get('primeSeat') in [True, 1, "1", "true", "True"] else 0, 
                         stall.get('primePriceINR'), 
                         stall.get('primePriceUSD')
                     ))

                 for am in layout_details.get('amenities', []):
                     cursor.execute("INSERT INTO stall_amenities (event_id, stall_name, amenity, qty) VALUES (%s, %s, %s, %s)", (event_id, am.get('stallName'), am.get('amenity'), am.get('qty')))

             # 4. Update Terms
             if terms_data:
                 cursor.execute("DELETE FROM event_terms WHERE event_id = %s", (event_id,))
                 for term in terms_data:
                     cursor.execute("INSERT INTO event_terms (event_id, policy_group, policy_type, policy_name, is_default) VALUES (%s,%s,%s,%s,%s)", (event_id, term.get('policyGroup'), term.get('policyType'), term.get('policyName'), term.get('isDefault', False)))

             # 5. Update Vendors
             if vendors_data:
                 cursor.execute("DELETE FROM event_vendors WHERE event_id = %s", (event_id,))
                 cursor.execute("DELETE FROM event_sponsors WHERE event_id = %s", (event_id,))
                 cursor.execute("DELETE FROM event_guests WHERE event_id = %s", (event_id,))
                 
                 for v in vendors_data.get('vendors', []):
                     cursor.execute("INSERT INTO event_vendors (event_id, vendor_type, vendor_name, pass_count) VALUES (%s,%s,%s,%s)", (event_id, v.get('vendorType'), v.get('vendorName'), v.get('passCount', 0)))
                 for s in vendors_data.get('sponsors', []):
                     cursor.execute("INSERT INTO event_sponsors (event_id, sponsor_name, sponsorship_type) VALUES (%s,%s,%s)", (event_id, s.get('sponsorName'), s.get('sponsorship')))
                 for g in vendors_data.get('guests', []):
                     cursor.execute("INSERT INTO event_guests (event_id, guest_name, designation, contact, image) VALUES (%s,%s,%s,%s,%s)", (event_id, g.get('name'), g.get('designation'), g.get('contact'), g.get('image')))

             # 6. Update Food & Vehicle Details
             cursor.execute("DELETE FROM event_food_items WHERE event_id = %s", (event_id,))
             for item in food_provision.get("items", []):
                 cursor.execute("""
                    INSERT INTO event_food_items (event_id, caterer_name, meal_type, food_type, price_inr, price_usd, menu_details)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                 """, (event_id, item.get("catererName"), item.get("mealType"), item.get("foodType"), item.get("priceINR") or 0, item.get("priceUSD") or 0, item.get("menuDetails")))
             
             cursor.execute("DELETE FROM event_vehicle_details WHERE event_id = %s", (event_id,))
             for item in vehicle_provision.get("details", []):
                 cursor.execute("""
                    INSERT INTO event_vehicle_details (event_id, vehicle_type, price_inr, price_usd)
                    VALUES (%s, %s, %s, %s)
                 """, (event_id, item.get("vehicleType"), item.get("priceINR") or 0, item.get("priceUSD") or 0))

             cursor.execute("DELETE FROM event_vehicle_addons WHERE event_id = %s", (event_id,))
             for item in vehicle_provision.get("addons", []):
                 cursor.execute("""
                    INSERT INTO event_vehicle_addons (event_id, is_parent, addon_name, price)
                    VALUES (%s, %s, %s, %s)
                 """, (event_id, item.get("isParent", False), item.get("addOnName"), item.get("price") or 0))

             # 7. Handle Banner
             delete_banner = request.form.get("delete_banner") == "true"
             if 'banner' in request.files:
                 banner = request.files['banner']
                 if banner:
                     cursor.execute("DELETE FROM event_files WHERE event_id = %s AND file_type = 'banner'", (event_id,))
                     folder, relative_cat, _ = get_file_folder(banner)
                     filename = secure_filename(banner.filename)
                     save_path = os.path.join(folder, filename)
                     db_path = f"/uploads/{relative_cat}/{filename}"
                     banner.save(save_path)
                     cursor.execute("INSERT INTO event_files (event_id, file_name, file_path, file_type) VALUES (%s, %s, %s, %s)", (event_id, filename, db_path, "banner"))
             elif delete_banner:
                 cursor.execute("DELETE FROM event_files WHERE event_id = %s AND file_type = 'banner'", (event_id,))

             # 8. Handle Documents (Identity Proofs)
             existing_doc_ids_str = request.form.get("existing_doc_ids", "[]")
             try:
                 existing_doc_ids = json.loads(existing_doc_ids_str)
             except Exception:
                 existing_doc_ids = []

             if existing_doc_ids:
                 format_strings = ','.join(['%s'] * len(existing_doc_ids))
                 query = f"DELETE FROM event_files WHERE event_id = %s AND file_type != 'banner' AND id NOT IN ({format_strings})"
                 cursor.execute(query, [event_id] + existing_doc_ids)
             else:
                 cursor.execute("DELETE FROM event_files WHERE event_id = %s AND file_type != 'banner'", (event_id,))

             doc_count = int(request.form.get("doc_count", 0))
             for i in range(doc_count):
                 file = request.files.get(f"docs_{i}")
                 if file:
                     folder, relative_cat, file_category = get_file_folder(file)
                     filename = secure_filename(file.filename)
                     save_path = os.path.join(folder, filename)
                     db_path = f"/uploads/{relative_cat}/{filename}"
                     file.save(save_path)
                     cursor.execute("""
                         INSERT INTO event_files (event_id, file_name, file_path, file_type, doc_type, doc_number)
                         VALUES (%s, %s, %s, %s, %s, %s)
                     """, (event_id, filename, db_path, file_category, request.form.get(f"doc_type_{i}"), request.form.get(f"doc_number_{i}")))

             conn.commit()
             return jsonify({"success": True, "message": "Event updated successfully"})
        
        return jsonify({"error": "No eventDetails provided"}), 400

    except Exception as e:
        conn.rollback()
        print(f"Update Error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route("/api/abstract", methods=["GET"])
def get_abstract():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT 
                event_code,
                event_name
            FROM event_details_table
            WHERE status = 'APPROVED'
        """)

        data = cursor.fetchall()

        return jsonify(data)

    except Exception as e:
        return jsonify({"error": str(e)})

    finally:
        cursor.close()
        conn.close()
@super_admin_bp.route("/api/event-passes", methods=["GET"])
def get_event_passes():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
           SELECT 
                id,
                event_code,
                event_name,
                start_date,
                end_date
            FROM event_details_table
            WHERE status = 'APPROVED'
        """)

        data = cursor.fetchall()

        return jsonify(data)

    except Exception as e:
        return jsonify({"error": str(e)})

    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route("/api/contacts", methods=["GET"])
def get_contacts():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM my_contacts ORDER BY created_at DESC")
        data = cursor.fetchall()
        
        return jsonify(data), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route("/api/contacts", methods=["POST"])
def add_contact():
    try:
        data = request.json
        print("Creating contact:", data)
        conn = get_db_connection()
        cursor = conn.cursor()

        query = """
            INSERT INTO my_contacts (
                name, email, mobile, user_type, group_name
            ) VALUES (%s, %s, %s, %s, %s)
        """
        
        cursor.execute(query, (
            data.get("name"),
            data.get("email"),
            data.get("mobile"),
            data.get("userType"),
            data.get("groupName")
        ))
        
        conn.commit()
        return jsonify({"message": "Contact created successfully"}), 201

    except Exception as e:
        print("Contact creation error:", str(e))
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route("/api/contacts/<int:contact_id>", methods=["PUT"])
def update_contact(contact_id):
    try:
        data = request.json
        conn = get_db_connection()
        cursor = conn.cursor()

        query = """
            UPDATE my_contacts 
            SET name=%s, email=%s, mobile=%s, user_type=%s, group_name=%s
            WHERE id=%s
        """
        cursor.execute(query, (
            data.get("name"),
            data.get("email"),
            data.get("mobile"),
            data.get("userType"),
            data.get("groupName"),
            contact_id
        ))
        
        conn.commit()
        return jsonify({"message": "Contact updated successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route("/api/contacts/<int:contact_id>", methods=["DELETE"])
def delete_contact(contact_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM my_contacts WHERE id=%s", (contact_id,))
        conn.commit()
        return jsonify({"message": "Contact deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@super_admin_bp.route("/api/exhibitor/bookings_details", methods=["GET"])
def get_exhibitor_bookings():
    try:
        exhibitor_id = request.args.get("exhibitor_id")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT
                id,
                company_name,
                CONCAT(first_name, ' ', last_name) AS name,
                mobile,
                email,
                stall_area,
                products,
                address,
                status
            FROM exhibitor_stall_bookings
            ORDER BY id DESC
        """)
        data = cursor.fetchall()
        return jsonify({"data": data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# -----------------------------------
# USER PROFILE ENDPOINTS
# -----------------------------------
# ─────────────────────────────────────────────────────────────
# GET USER PROFILE
# ─────────────────────────────────────────────────────────────
@super_admin_bp.route("/api/user/profile/<int:user_id>", methods=["GET"])
def get_profile(user_id):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT id, name, email, mobile, address, country, state, city, profile_image, organization_name
            FROM users
            WHERE id = %s
        """, (user_id,))

        user = cursor.fetchone()

        if not user:
            return jsonify({
                "status": "error",
                "message": "User not found"
            }), 404

        return jsonify({
            "status": "success",
            "data": user
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# ─────────────────────────────────────────────────────────────
# UPDATE USER PROFILE
# ─────────────────────────────────────────────────────────────
@super_admin_bp.route("/api/user/update_profile", methods=["POST"])
def update_profile():
    conn = None
    cursor = None
    try:
        data = request.json
        user_id = data.get("id")

        if not user_id:
            return jsonify({
                "status": "error",
                "message": "User ID missing"
            }), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        print(f"Updating profile for user_id: {user_id}")
        
        cursor.execute("""
            UPDATE users
            SET name=%s,
                mobile=%s,
                address=%s,
                country=%s,
                state=%s,
                city=%s,
                profile_image=%s,
                organization_name=%s
            WHERE id=%s
        """, (
            data.get("name"),
            data.get("mobile"),
            data.get("address"),
            data.get("country"),
            data.get("state"),
            data.get("city"),
            data.get("profile_image"),
            data.get("organization_name"),
            user_id
        ))

        rows_affected = cursor.rowcount
        print(f"Rows affected: {rows_affected}")

        conn.commit()

        if rows_affected == 0:
            # Check if user exists to distinguish between "no change" and "user not found"
            cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
            if not cursor.fetchone():
                return jsonify({
                    "status": "error",
                    "message": "User not found"
                }), 404
            
            return jsonify({
                "status": "success",
                "message": "Profile is already up to date"
            })

        return jsonify({
            "status": "success",
            "message": "Profile updated successfully"
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def get_next_complaint_code():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(id) FROM complaint")
        result = cursor.fetchone()
        next_id = (result[0] or 0) + 1
        return f"CMP-{str(next_id).zfill(4)}"
    except Exception as e:
        print("Error generating complaint code:", e)
        return "CMP-0001"
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@super_admin_bp.route('/api/complaints', methods=['GET'])
def get_complaints():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM complaint ORDER BY created_on DESC")
        data = cursor.fetchall()
        for row in data:
            if row.get('created_on'):
                row['created_on'] = row['created_on'].strftime('%Y-%m-%d %H:%M')
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@super_admin_bp.route('/api/approved-events', methods=['GET'])
def get_approved_events():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, event_name FROM event_details_table WHERE status = 'APPROVED'")
        data = cursor.fetchall()
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# Create new complaint
@super_admin_bp.route('/api/complaints', methods=['POST'])
def create_complaint():
    conn = None
    cursor = None
    try:
        data = request.json
        if not data or not data.get('event_id'):
            return jsonify({'error': 'Event is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get event name
        cursor.execute("SELECT event_name FROM event_details_table WHERE id = %s", (data['event_id'],))
        event_result = cursor.fetchone()
        
        if not event_result:
            return jsonify({'error': 'Invalid event'}), 400
        
        event_name = event_result['event_name']
        complaint_code = get_next_complaint_code()
        
        # Insert complaint
        query = """
            INSERT INTO complaint (
                complaint_code, event_id, event_name,
                infrastructure_rating, amenities_rating, overall_experience_rating,
                venue_locations_rating, transportation_rating, convenience_rating,
                explanation, status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        params = (
            complaint_code,
            data['event_id'],
            event_name,
            data.get('infrastructure_rating', 0),
            data.get('amenities_rating', 0),
            data.get('overall_experience_rating', 0),
            data.get('venue_locations_rating', 0),
            data.get('transportation_rating', 0),
            data.get('convenience_rating', 0),
            data.get('explanation', ''),
            'Active'
        )
        
        cursor.execute(query, params)
        conn.commit()
        complaint_id = cursor.lastrowid
        
        return jsonify({
            'success': True,
            'id': complaint_id,
            'complaint_code': complaint_code
        }), 201

    except Exception as e:
        print("CREATE COMPLAINT ERROR:", str(e))
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
@super_admin_bp.route("/api/programs", methods=["POST"])
def create_program():
    try:
        data = request.json
        print("Creating program:", data)
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Generate unique Program Code
        cursor.execute("SELECT MAX(id) as max_id FROM event_programs")
        res = cursor.fetchone()
        next_id = (res['max_id'] or 0) + 1
        program_code = f"PRG-{str(next_id).zfill(3)}"

        query = """
            INSERT INTO event_programs (
                event_id, program_name, program_code, category, type, 
                start_date, end_date, venue, max_participants, budget, 
                coordinator_name, coordinator_email, description, status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        # Handle empty date strings by converting to None correctly 
        start_date = data.get("start") if data.get("start") else None
        end_date = data.get("end") if data.get("end") else None
        max_part = data.get("maxPart") if data.get("maxPart") else 0
        budget = data.get("budget") if data.get("budget") else 0.0

        cursor.execute(query, (
            data.get("event_id"),
            data.get("name"),
            program_code,
            data.get("category"),
            data.get("type"),
            start_date,
            end_date,
            data.get("venue"),
            max_part,
            budget,
            data.get("coordName"),
            data.get("coordEmail"),
            data.get("desc"),
            data.get("status", "Active")
        ))
        
        conn.commit()
        return jsonify({"message": "Program created successfully", "program_code": program_code}), 201

    except Exception as e:
        print("Program creation error:", str(e))
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@super_admin_bp.route('/api/program-events', methods=['GET'])
def get_program_events():
    conn = None
    cursor = None
    try:
        organizer_id = request.args.get("organizer_id")
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        # Fetch events along with program counts by status
        query = """
            SELECT 
                e.id, 
                e.event_code, 
                e.event_name,
                e.start_date,
                e.end_date,
                COUNT(CASE WHEN p.status = 'Active' THEN 1 END) as approved,
                COUNT(CASE WHEN p.status = 'Inactive' THEN 1 END) as rejected,
                COUNT(CASE WHEN p.status = 'Inprocess' THEN 1 END) as inprocess,
                COUNT(p.id) as created
            FROM event_details_table e
            LEFT JOIN event_programs p ON e.id = p.event_id
            WHERE e.status = 'APPROVED'
        """
        
        if organizer_id:
            query += " AND e.user_id = %s "
            query += " GROUP BY e.id ORDER BY e.id DESC"
            cursor.execute(query, (organizer_id,))
        else:
            query += " GROUP BY e.id ORDER BY e.id DESC"
            cursor.execute(query)
        data = cursor.fetchall()

        # Format dates for JSON
        for row in data:
            if row.get("start_date"):
                row["start_date"] = row["start_date"].strftime("%Y-%m-%d")
            if row.get("end_date"):
                row["end_date"] = row["end_date"].strftime("%Y-%m-%d")

        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@super_admin_bp.route('/api/program-list/<int:event_id>', methods=['GET'])
def get_programs_by_event(event_id):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM event_programs WHERE event_id = %s ORDER BY id DESC", (event_id,))
        data = cursor.fetchall()
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@super_admin_bp.route('/api/complaints/<int:complaint_id>', methods=['DELETE'])
def delete_complaint(complaint_id):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM complaint WHERE id = %s", (complaint_id,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Complaint deleted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ── FEEDBACK ROUTES ──────────────────────────────────────────────────────────

# Get next feedback code
def get_next_feedback_code():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute("SELECT feedback_code FROM feedback_event ORDER BY id DESC LIMIT 1")
        result = cursor.fetchone()
        if result:
            last_code = result['feedback_code']
            try:
                # Extract number from FBK-X
                parts = last_code.split('-')
                if len(parts) > 1:
                    number = int(parts[1])
                    return f"FBK-{number + 1}"
            except (ValueError, IndexError):
                pass
        return "FBK-1"
    finally:
        cursor.close()
        db.close()

# Get all feedbacks
@super_admin_bp.route('/api/feedbacks', methods=['GET'])
def get_feedbacks():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        query = """
            SELECT
                id, feedback_code, event_id, event_name,
                status, explanation, created_at, modified_on
            FROM feedback_event
            ORDER BY id DESC
        """
        cursor.execute(query)
        feedbacks = cursor.fetchall()
        return jsonify(feedbacks or [])
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        db.close()

# Get single feedback by ID
@super_admin_bp.route('/api/feedbacks/<int:feedback_id>', methods=['GET'])
def get_feedback(feedback_id):
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM feedback_event WHERE id = %s", (feedback_id,))
        result = cursor.fetchone()
        if not result:
            return jsonify({'error': 'Feedback not found'}), 404
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        db.close()

# Create new feedback
@super_admin_bp.route('/api/feedbacks', methods=['POST'])
def create_feedback():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        data = request.json
        if not data.get('event_id'):
            return jsonify({'error': 'Event is required'}), 400

        # Get event name
        cursor.execute("SELECT event_name FROM event_details_table WHERE id = %s", (data['event_id'],))
        event_result = cursor.fetchone()
        if not event_result:
            return jsonify({'error': 'Invalid event'}), 400

        event_name = event_result['event_name']
        feedback_code = get_next_feedback_code()

        query = """
            INSERT INTO feedback_event (
                feedback_code, event_id, event_name,
                explanation, status
            ) VALUES (%s, %s, %s, %s, 'Active')
        """
        params = (feedback_code, data['event_id'], event_name, data.get('explanation', ''))
        cursor.execute(query, params)
        db.commit()
        feedback_id = cursor.lastrowid

        return jsonify({
            'success': True,
            'id': feedback_id,
            'feedback_code': feedback_code
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        db.close()

# Update existing feedback
@super_admin_bp.route('/api/feedbacks/<int:feedback_id>', methods=['PUT'])
def update_feedback(feedback_id):
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        data = request.json
        if not data.get('event_id'):
            return jsonify({'error': 'Event is required'}), 400

        # Get event name
        cursor.execute("SELECT event_name FROM event_details_table WHERE id = %s", (data['event_id'],))
        event_result = cursor.fetchone()
        if not event_result:
            return jsonify({'error': 'Invalid event'}), 400

        event_name = event_result['event_name']

        query = """
            UPDATE feedback_event
            SET event_id    = %s,
                event_name  = %s,
                explanation = %s,
                modified_on = CURDATE()
            WHERE id = %s
        """
        params = (data['event_id'], event_name, data.get('explanation', ''), feedback_id)
        cursor.execute(query, params)
        db.commit()
        return jsonify({'success': True, 'message': 'Feedback updated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        db.close()

# Delete feedback
@super_admin_bp.route('/api/feedbacks/<int:feedback_id>', methods=['DELETE'])
def delete_feedback(feedback_id):
    db = get_db_connection()
    cursor = db.cursor()
    try:
        cursor.execute("DELETE FROM feedback_event WHERE id = %s", (feedback_id,))
        db.commit()
        return jsonify({'success': True, 'message': 'Feedback deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        db.close()
@super_admin_bp.route("/api/policies/export/excel", methods=["GET"])
def export_policies_excel():
    try:
        organizer_id = request.args.get('organizer_id')
        print(f"INFO: [{datetime.now()}] Initiating Excel export for Policy Management. Organizer ID: {organizer_id}")
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        query = "SELECT policy_code, policy_name, policy_type, policy_group, description, status FROM policies"
        params = []
        if organizer_id:
            query += " WHERE organizer_id = %s"
            params.append(organizer_id)
        query += " ORDER BY id DESC"
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        
        if not data:
            return jsonify({"error": "No data available for export"}), 404

        df = pd.DataFrame(data)
        df.columns = ["Policy Code", "Policy Name", "Policy Type", "Policy Group", "Description", "Status"]
        
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Policy_Report')
            workbook = writer.book
            worksheet = writer.sheets['Policy_Report']

            # Header Styling
            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")

            # Alternate Row Styling & Column Width
            thin_border = openpyxl.styles.Side(border_style="thin", color="000000")
            border = openpyxl.styles.Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(data)+1), start=2):
                fill = openpyxl.styles.PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") if row_idx % 2 == 0 else None
                for cell in row:
                    if fill: cell.fill = fill
                    cell.border = border

            # Auto-adjust columns width
            for idx, col in enumerate(df.columns):
                try:
                    max_len = max(df[col].astype(str).map(len).max(), len(col)) + 4
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = min(max_len, 50)
                except:
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = 20

            worksheet.freeze_panes = 'A2'

        output.seek(0)
        filename = f"policy_list_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
        print(f"INFO: [{datetime.now()}] Excel export completed: {filename}")
        
        # ✅ Safest alternative: Return raw bytes with explicit headers
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response

    except Exception as e:
        print(f"ERROR: [{datetime.now()}] Excel export failed: {str(e)}")
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@super_admin_bp.route("/api/policies/export/pdf", methods=["GET"])
def export_policies_pdf():
    try:
        organizer_id = request.args.get('organizer_id')
        print(f"INFO: [{datetime.now()}] Initiating PDF export for Policy Management. Organizer ID: {organizer_id}")
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        query = "SELECT policy_code, policy_name, policy_type, policy_group, status FROM policies"
        params = []
        if organizer_id:
            query += " WHERE organizer_id = %s"
            params.append(organizer_id)
        query += " ORDER BY id DESC"
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        
        if not data:
            return jsonify({"error": "No data available for export"}), 404

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=30)
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Title Section
        title_style = styles['Title']
        title_style.fontSize = 20
        title_style.textColor = colors.HexColor("#1e40af")
        elements.append(Paragraph("Policy Management Report", title_style))
        
        date_style = styles['Normal']
        date_style.alignment = 1
        elements.append(Paragraph(f"Exported on: {datetime.now().strftime('%B %d, %Y | %H:%M:%S')}", date_style))
        elements.append(Spacer(1, 24))

        # Table Construction
        table_data = [["Policy Code", "Policy Name", "Type", "Group", "Status"]]
        
        for row in data:
            table_data.append([
                row['policy_code'],
                row['policy_name'],
                row['policy_type'],
                row['policy_group'],
                row['status']
            ])

        # Table Formatting
        column_widths = [100, 200, 100, 150, 80]
        t = Table(table_data, colWidths=column_widths, repeatRows=1)
        
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(t)
        doc.build(elements)
        
        output.seek(0)
        filename = f"policy_list_{datetime.now().strftime('%Y_%m_%d')}.pdf"
        print(f"INFO: [{datetime.now()}] PDF export completed: {filename}")
        
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')

    except Exception as e:
        print(f"ERROR: [{datetime.now()}] PDF export failed: {str(e)}")
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@super_admin_bp.route("/api/venues/export/excel", methods=["GET"])
def export_venues_excel():
    try:
        print(f"INFO: [{datetime.now()}] Initiating Excel export for Venues.")
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT venue_code, venue_name, address, status FROM venues ORDER BY id DESC")
        data = cursor.fetchall()
        
        if not data:
            return jsonify({"error": "No data available for export"}), 404

        df = pd.DataFrame(data)
        df.columns = ["Venue Code", "Venue Name", "Address", "Status"]
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Venue_Report')
            
            workbook = writer.book
            worksheet = writer.sheets['Venue_Report']
            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")
            
            thin_border = openpyxl.styles.Side(border_style="thin", color="000000")
            border = openpyxl.styles.Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
            
            for row in worksheet.iter_rows(min_row=2, max_row=len(data)+1):
                for cell in row:
                    cell.border = border

            for idx, col in enumerate(df.columns):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = 20

        output.seek(0)
        filename = f"venue_list_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
        
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response
    except Exception as e:
        print(f"ERROR: [{datetime.now()}] Excel export failed: {str(e)}")
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@super_admin_bp.route("/api/venues/export/pdf", methods=["GET"])
def export_venues_pdf():
    try:
        print(f"INFO: [{datetime.now()}] Initiating PDF export for Venues.")
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT venue_code, venue_name, address, status FROM venues ORDER BY id DESC")
        data = cursor.fetchall()

        if not data:
            return jsonify({"error": "No data available for export"}), 404
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = styles['Title']
        title_style.fontSize = 20
        title_style.textColor = colors.HexColor("#1e40af")
        elements.append(Paragraph("Venue Management Report", title_style))
        
        date_style = styles['Normal']
        date_style.alignment = 1
        elements.append(Paragraph(f"Exported on: {datetime.now().strftime('%B %d, %Y | %H:%M:%S')}", date_style))
        elements.append(Spacer(1, 24))

        table_data = [["Venue Code", "Venue Name", "Address", "Status"]]
        for row in data:
            table_data.append([
                row['venue_code'],
                row['venue_name'],
                row['address'],
                row['status']
            ])

        column_widths = [100, 200, 250, 100]
        t = Table(table_data, colWidths=column_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(t)
        doc.build(elements)
        
        output.seek(0)
        filename = f"venue_list_{datetime.now().strftime('%Y_%m_%d')}.pdf"
        print(f"INFO: [{datetime.now()}] PDF export completed: {filename}")
        
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')
    except Exception as e:
        print(f"ERROR: [{datetime.now()}] PDF export failed: {str(e)}")
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()
